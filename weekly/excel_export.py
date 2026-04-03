#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Weekly-specific Excel/PDF export για ενημερωτικά ΣΗΘΥΑ.
"""

import logging
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from MONTHLY.config import BASE_DIR, MAX_FILENAME_CHARS, MAX_FOLDER_CHARS
from MONTHLY.helpers import clipped_folder_name, sanitize_name

log = logging.getLogger("weekly.excel_export")

TEMPLATE_FILE = Path(__file__).resolve().parent / "WEEKLY_Invoice_GREEN_VALUE_01.xlsx"
OUTPUT_BASE_DIR = BASE_DIR / "ΕΝΗΜΕΡΩΤΙΚΑ_ΣΗΜΕΙΩΜΑΤΑ_ΕΒΔΟΜΑΔΙΑΙΑ"


def clipped_filename_weekly(company_name: str, tag: str, ext: str, limit=MAX_FILENAME_CHARS):
    prefix = "ΕΒΔΟΜΑΔΙΑΙΟ_ΣΗΜΕΙΩΜΑ_"
    base = sanitize_name(company_name).replace(" ", "_")
    cand = f"{prefix}{base}_{tag}.{ext}"
    if len(cand) <= limit:
        return cand
    over = len(cand) - limit
    base_cut = base[:max(1, len(base) - over)]
    cand = f"{prefix}{base_cut}_{tag}.{ext}"
    if len(cand) > limit:
        trunk = f"{prefix}{tag}"
        cand = trunk[:limit - (len(ext) + 1)] + "." + ext
    return cand


def xlsx_filename_weekly(company_name, tag):
    return clipped_filename_weekly(company_name, tag, "xlsx", MAX_FILENAME_CHARS)


def pdf_filename_weekly(company_name, tag):
    return clipped_filename_weekly(company_name, tag, "pdf", MAX_FILENAME_CHARS)


def make_week_dirs(start_date: pd.Timestamp, end_date: pd.Timestamp):
    iso_year, iso_week, _ = start_date.isocalendar()
    tag = f"{iso_year}-W{iso_week:02d}"
    root = OUTPUT_BASE_DIR / tag
    xlsx_dir = root / "XLSX"
    pdf_dir = root / "PDF"
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)
    return tag, root, xlsx_dir, pdf_dir


def determine_pdf_subfolder_name(email, email_to_companies):
    companies = email_to_companies.get(email, [])
    return clipped_folder_name([], companies, limit=MAX_FOLDER_CHARS)


def set_cell_value(ws: Worksheet, coord: str, value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                min_col, min_row = merged_range.bounds[0], merged_range.bounds[1]
                ws.cell(row=min_row, column=min_col).value = value
                return
    ws[coord].value = value


def set_cell_property(ws: Worksheet, coord: str, prop_name: str, prop_value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                min_col, min_row = merged_range.bounds[0], merged_range.bounds[1]
                cell = ws.cell(row=min_row, column=min_col)
                break
    setattr(cell, prop_name, prop_value)


def _ranges_intersect(left: str, right: str) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = range_boundaries(left)
    right_min_col, right_min_row, right_max_col, right_max_row = range_boundaries(right)
    return not (
        left_max_col < right_min_col or
        right_max_col < left_min_col or
        left_max_row < right_min_row or
        right_max_row < left_min_row
    )


def _unmerge_overlaps(ws: Worksheet, coord: str):
    for merged_range in list(ws.merged_cells.ranges):
        if _ranges_intersect(str(merged_range), coord):
            try:
                ws.unmerge_cells(str(merged_range))
            except Exception as exc:
                log.warning("unmerge %s failed before merge %s: %s", merged_range, coord, exc)


def _safe_merge(ws, coord: str):
    try:
        _unmerge_overlaps(ws, coord)
        ws.merge_cells(coord)
        return True
    except Exception as exc:
        log.warning("merge %s failed: %s", coord, exc)
        return False


def _add_logo_if_available(ws, base_dir: Path):
    try:
        if getattr(ws, "_images", []):
            return
    except Exception:
        pass

    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        return

    logo_candidates = [base_dir / "LOGO.png", base_dir / "logo.png"]
    logo_path = next((path for path in logo_candidates if path.exists()), None)
    if logo_path is None:
        return

    try:
        img = XLImage(str(logo_path))
        img.anchor = "A1"
        img.width = 120
        img.height = 60
        ws.add_image(img)
    except Exception as exc:
        log.warning("logo add failed: %s", exc)


def generate_invoice_excel_weekly(df_daily, summary, producer_row, start_date, end_date, xlsx_output_dir: Path, tag: str):
    try:
        if not TEMPLATE_FILE.exists():
            raise FileNotFoundError(f"Λείπει template: {TEMPLATE_FILE}")

        company_name = str(producer_row["Εταιρεία"].values[0])
        email_value = str(producer_row["Email"].values[0]) if "Email" in producer_row else ""
        iban = producer_row["IBAN"].values[0] if "IBAN" in producer_row else ""
        rate = float(producer_row["Μοναδιαία Χρέωση ΦοΣΕ"].values[0])
        _, _, sum_prov = summary

        out_name = xlsx_filename_weekly(company_name, tag)
        xlsx_path = xlsx_output_dir / out_name
        xlsx_output_dir.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active

        try:
            ws.print_area = "A1:H55"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass

        _add_logo_if_available(ws, BASE_DIR)
        for coord in ("C2", "D2", "E2", "F2", "G2"):
            try:
                ws[coord].value = None
            except Exception:
                pass

        _safe_merge(ws, "D1:F2")
        ws["D1"].value = (
            "Ενημερωτικό Σημείωμα Εβδομάδας\n"
            f"{start_date.strftime('%d/%m/%y')} – {end_date.strftime('%d/%m/%y')}"
        )
        ws["D1"].font = Font(bold=True, size=14)
        ws["D1"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

        _safe_merge(ws, "G1:H1")
        ws["G1"].value = (
            "Φορέας Σωρευτικής Εκπροσώπησης ΑΠΕ (Φο.Σ.Ε.)\n"
            "Διεύθυνση: Φιλοπάππου 19, Αθήνα 11741, Ελλάδα\n"
            "ΑΦΜ: 801961185\n"
            "ΓΕΜΗ: 167104201000\n"
            "ΔΟΥ:ΦΑΕ Αθηνών\n"
            "Email: info@greenvalue.gr"
        )
        ws["G1"].font = Font(bold=False, size=10)
        ws["G1"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        needed = ["Α.Μ. ΑΠΕ", "Εταιρεία", "ΑΦΜ", "ΔΟΥ", "Διεύθυνση", "Email", "Τεχνολογία"]
        vals = producer_row.iloc[0][needed].tolist()
        for cell_ref, val in zip(["B4", "C4", "D4", "E4", "F4", "G4", "H4"], vals):
            set_cell_value(ws, cell_ref, val)
        set_cell_property(ws, "C4", "font", Font(size=13))

        set_cell_value(ws, "B6", pd.Timestamp.today().strftime("%d/%m/%y"))
        _safe_merge(ws, "D6:F6")
        ws["D6"].value = f"{start_date.strftime('%d/%m/%y')} - {end_date.strftime('%d/%m/%y')}"
        ws["D6"].font = Font(bold=True, size=14)
        ws["D6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for merged_range in list(ws.merged_cells.ranges):
            bounds = merged_range.bounds
            if bounds[1] >= 10 and bounds[1] <= 19 and bounds[0] >= 3 and bounds[0] <= 7:
                try:
                    ws.unmerge_cells(merged_range.coord)
                except Exception:
                    pass

        for row in ws.iter_rows(min_row=10, max_row=19, min_col=3, max_col=7):
            for cell in row:
                try:
                    cell.value = None
                except Exception:
                    pass

        start_row = 10
        for r_idx, row_vals in enumerate(df_daily.values, start=start_row):
            for c_idx, value in enumerate(row_vals, start=3):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"

        total_row = start_row + len(df_daily) - 1
        for col in range(3, 8):
            set_cell_property(ws, f"{chr(64 + col)}{total_row}", "font", Font(bold=True, size=15))

        _safe_merge(ws, "C28:D28")
        ws["C28"].value = iban
        ws["C28"].font = Font(bold=True, size=14)
        ws["C28"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        _safe_merge(ws, "C29:D29")
        ws["C29"].value = (pd.Timestamp.today() + pd.Timedelta(days=5)).strftime("%d/%m/%y")
        ws["C29"].font = Font(bold=True, size=14)
        ws["C29"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        set_cell_value(ws, "D21", rate)
        set_cell_value(ws, "D22", round(sum_prov, 2))
        set_cell_property(ws, "D43", "number_format", "#,##0.00")

        wb.save(xlsx_path)
        print(f"✅ XLSX → {xlsx_path}")
        return str(xlsx_path), company_name, email_value

    except Exception as exc:
        log.error("Excel generation failed: %s", exc)
        print(f"❌ Excel generation failed: {exc}")
        return None, None, None


def _verify_pdf(path: str, min_bytes: int = 500):
    return os.path.exists(path) and os.path.getsize(path) >= min_bytes


def export_to_pdf_with_excel(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    try:
        import xlwings as xw
    except Exception:
        return False, "excel-not-available"

    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(os.path.abspath(xlsx_path))
        sht = wb.sheets.active
        try:
            ps = sht.api.PageSetup
            ps.Zoom = False
            ps.FitToPagesWide = 1
            ps.FitToPagesTall = 1
            ps.PrintArea = "A1:H55"
        except Exception:
            pass

        out_pdf = os.path.abspath(pdf_path)
        Path(os.path.dirname(out_pdf)).mkdir(parents=True, exist_ok=True)

        if hasattr(wb, "to_pdf"):
            wb.to_pdf(out_pdf)
        else:
            raise AttributeError("wb.to_pdf not available")

        try:
            wb.close()
        except Exception:
            pass

        if _verify_pdf(out_pdf):
            return True, "excel"
        return False, "excel-empty"

    except Exception as exc:
        log.error("PDF export failed: %s", exc)
        return False, f"excel-error:{exc}"
    finally:
        try:
            app.quit()
        except Exception:
            pass


def export_to_pdf(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    ok, how = export_to_pdf_with_excel(xlsx_path, pdf_path)
    if ok:
        return True, how
    return False, how
