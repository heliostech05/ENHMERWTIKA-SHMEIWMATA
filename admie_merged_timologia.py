#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
enimerwtika_simeiwmata_new.py  (15λεπτο + mac-friendly + mixed granularity + DST-safe)

ΤΙ ΚΑΝΕΙ
---------
- Αγνοεί πλήρως DAM γραμμές πριν από 2025-10-01 (καθώς είναι ωριαίες).
- Από 2025-10-01 και μετά: διαβάζει DAM 15' (START TIME) όπως δίνεται στο αρχείο,
  ΧΩΡΙΣ επιπλέον -15 λεπτά (δηλ. δεν το μετακινούμε).
- Κρατά την **αρχική σειρά** του αρχείου DAM (ΔΕΝ κάνει sort), ώστε στις 26/10 να εμφανίζεται
  ακριβώς: 02:45, 03:00, 03:15, 03:30, 03:45, 03:00, 03:15, 03:30, 03:45, 04:00, ...
- Χειρίζεται τη διπλή ώρα της 26/10/2025 (DST fall-back) με index/φίλτρα
  και κάνει pairing παραγωγής–DAM 1-προς-1 με βάση index.
- Υπολογίζει κόστος ανά 15' και ημερήσια σύνολα.
- Παράγει Excel από template και (αν υπάρχει MS Excel) κάνει export σε PDF με xlwings (Mac/Win).

ΑΠΑΙΤΗΣΕΙΣ
----------
pip install pandas openpyxl xlwings
"""

import os
import re
import pandas as pd
import shutil
import subprocess
import time
from collections import defaultdict
from pathlib import Path
import matplotlib
matplotlib.use("Agg")
matplotlib.rcParams['font.family'] = 'Century Gothic'
from matplotlib import font_manager as fm
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import openpyxl.styles as xls

# DEBUG: ποια μέρα θέλουμε να τυπώσουμε αναλυτικά;
# Παράδειγμα: "2025-10-10". Βάλε None για να το απενεργοποιήσεις.
DEBUG_DAY = "2025-10-01"

# ===== Optional: xlwings (Excel PDF export). No pywin32 needed on macOS =====
try:
    import xlwings as xw
    _HAS_XLWINGS = True
except Exception:
    _HAS_XLWINGS = False

# ===== Simple logging =====
LOG_BASE = "logs/timologia"
os.makedirs(LOG_BASE, exist_ok=True)
def log(name, msg):
    with open(os.path.join(LOG_BASE, f"{name}.txt"), "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")

# ===== Path/name limits =====
MAX_FOLDER_CHARS   = 120
MAX_FILENAME_CHARS = 140

# Register local font files explicitly (works without system install)
font_dir = os.path.join(os.path.dirname(__file__), "fonts")
for fname in ["CenturyGothic.ttf", "GOTHICB.ttf", "GOTHICI.ttf", "GOTHICBI.ttf"]:
    fpath = os.path.join(font_dir, fname)
    if os.path.exists(fpath):
        fm.fontManager.addfont(fpath)

# Tell Matplotlib to use this family name (try both common namings)
matplotlib.rcParams["font.family"] = "Century Gothic"
matplotlib.rcParams["font.sans-serif"] = ["Century Gothic", "CenturyGothic", "DejaVu Sans"]

# ===== Helpers =====
WIN_RESERVED = {"CON","PRN","AUX","NUL","COM1","COM2","COM3","COM4","COM5","COM6","COM7","COM8","COM9",
                "LPT1","LPT2","LPT3","LPT4","LPT5","LPT6","LPT7","LPT8","LPT9"}

def sanitize_name(name: str) -> str:
    if name is None:
        return "UNTITLED"
    s = str(name)
    s = re.sub(r'[\\/*?:"<>|]', "", s)
    s = s.strip().rstrip(".")
    if not s:
        s = "UNTITLED"
    if s.upper() in WIN_RESERVED:
        s = "_" + s
    return s

def normalize_name(name: str) -> str:
    return re.sub(r'[\s._\\-]', '', str(name).strip().lower())

def join_with_limit(parts, sep=" & ", limit=120):
    out, total = [], 0
    for p in parts:
        piece = p if not out else (sep + p)
        if total + len(piece) > limit:
            break
        out.append(p)
        total += len(piece)
    if out:
        return sep.join(out)
    return (parts[0] if parts else "UNTITLED")[:limit]

def clipped_folder_name(preferred_names, fallback_names, limit=MAX_FOLDER_CHARS):
    def build_name(items):
        uniq = sorted({sanitize_name(x) for x in items if x})
        return join_with_limit(uniq, sep=" & ", limit=limit)
    if preferred_names:
        return build_name(preferred_names)
    return build_name(fallback_names)

def clipped_filename(company_name: str, month: str, ext: str, limit=MAX_FILENAME_CHARS):
    prefix = "ΕΝΗΜΕΡΩΤΙΚΟ_ΣΗΜΕΙΩΜΑ_"
    base = sanitize_name(company_name).replace(" ", "_")
    cand = f"{prefix}{base}_{month}.{ext}"
    if len(cand) <= limit:
        return cand
    over = len(cand) - limit
    base_cut = base[:max(1, len(base) - over)]
    cand = f"{prefix}{base_cut}_{month}.{ext}"
    if len(cand) > limit:
        trunk = f"{prefix}{month}"
        cand = trunk[:limit - (len(ext) + 1)] + "." + ext
    return cand

def _add_daily_plot(ws, df_daily, anchor_cell="B56", color_hex="#22A052"):
    """
    Barplot ημερήσιας παραγωγής, χωρίς να περιλαμβάνει τη γραμμή 'Σύνολο'.
    """
    try:
        df_plot = df_daily[df_daily["Περίοδος εκκαθάρισης"].str.contains("Σύνολο", case=False) == False].copy()
        x = df_plot["Περίοδος εκκαθάρισης"].astype(str)
        y = pd.to_numeric(df_plot["ΕΝΕΡΓΕΙΑ (kWh)"], errors="coerce")

        plt.figure(figsize=(7.5, 3.0), dpi=160)
        plt.bar(x, y, color=color_hex, width=0.6)

        plt.title("Διάγραμμα Ημερήσιας Παραγωγής (kWh)", fontname="Century Gothic", fontsize=13, pad=15)
        plt.xlabel("")
        plt.ylabel("kWh", fontname="Century Gothic", fontsize=11)
        plt.xticks(rotation=45, ha="right", fontsize=9, fontname="Century Gothic")
        plt.yticks(fontsize=9, fontname="Century Gothic")
        plt.grid(True, linestyle="--", alpha=0.4)
        plt.tight_layout(pad=1.0)

        plot_path = "daily_prod_plot.png"
        plt.savefig(plot_path, dpi=160, transparent=True)
        plt.close()

        img = XLImage(plot_path)
        img.anchor = anchor_cell
        img.width = 1100
        img.height = 400
        ws.add_image(img)

    except Exception as e:
        print(f"⚠️ Plot insertion failed: {e}")

def fix_header_rows(ws):
    ws.row_dimensions[1].height = 160
    ws.row_dimensions[2].height = 80
    ws.row_dimensions[4].height = 100
    ws.row_dimensions[58].height = 80

    ws["C2"].alignment = xls.Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=False
    )

# ===== Producers / Email groups =====
def load_producers(filepath="producers.xlsx"):
    fn = "load_producers"
    try:
        df = pd.read_excel(filepath, dtype={'Code': str})
        if 'Εταιρεία' not in df.columns:
            raise ValueError("Λείπει στήλη 'Εταιρεία'")
        if 'Email' not in df.columns:
            df['Email'] = ""
        if 'Μοναδιαία Χρέωση ΦοΣΕ' not in df.columns:
            raise ValueError("Λείπει στήλη 'Μοναδιαία Χρέωση ΦοΣΕ'")
        if 'Όνομα Φακέλου' not in df.columns:
            df['Όνομα Φακέλου'] = ""
        df['normalized_name'] = df['Εταιρεία'].astype(str).apply(normalize_name)
        log(fn, f"OK producers: {len(df)}")
        return df
    except Exception as e:
        log(fn, f"ERROR {e}")
        return None

def build_email_groups(producers_df):
    email_to_companies = defaultdict(set)
    email_to_customs   = defaultdict(set)
    for _, row in producers_df.iterrows():
        email = (str(row.get('Email', '') or '').strip()) or "NO_EMAIL"
        comp  = str(row.get('Εταιρεία', '') or '').strip()
        cust  = str(row.get('Όνομα Φακέλου', '') or '').strip()
        if comp:
            email_to_companies[email].add(comp)
        if cust:
            email_to_customs[email].add(cust)
    email_to_companies = {em: sorted(v) for em, v in email_to_companies.items()}
    email_to_customs   = {em: sorted(v) for em, v in email_to_customs.items()}
    return email_to_companies, email_to_customs

# ===== DAM 15' (START) with disclaimer header detection & DST-safe duplicates =====

HEADER_TS_KEYS = ["date", "time", "timestamp", "cet", "ce(s)t", "gmt", "utc", "eet", "athens", "gmt+2"]
HEADER_PRICE_KEYS = ["price", "eur/mwh", "€/mwh", "auction", "day-ahead", "day ahead"]

def _find_header_line(path, max_scan=200):
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        for i in range(max_scan):
            line = f.readline()
            if not line:
                break
            low = line.strip().lower()
            if any(k in low for k in HEADER_TS_KEYS) and any(k in low for k in HEADER_PRICE_KEYS):
                return i
    return 1

def _infer_dam_columns(df: pd.DataFrame):
    cols = list(df.columns)
    lower = {c: c.lower() for c in cols}

    ts_cands = [c for c in cols if any(k in lower[c] for k in HEADER_TS_KEYS)]
    price_cands = [c for c in cols if any(k in lower[c] for k in HEADER_PRICE_KEYS)]

    ts_col = ts_cands[0] if ts_cands else None
    price_cands = [c for c in price_cands if c != ts_col]
    price_col = price_cands[0] if price_cands else None

    if not ts_col:
        best, best_rate = None, -1
        for c in cols:
            try:
                parsed = pd.to_datetime(df[c], errors="coerce")
                rate = parsed.notna().sum() / max(1, df[c].notna().sum())
                if rate >= 0.8 and rate > best_rate:
                    best, best_rate = c, rate
            except Exception:
                pass
        ts_col = best
    if not price_col:
        best, best_rate = None, -1
        for c in cols:
            if c == ts_col:
                continue
            s = pd.to_numeric(df[c].astype(str).str.replace(",", ".", regex=False), errors="coerce")
            rate = s.notna().sum() / max(1, df[c].notna().sum())
            if rate >= 0.8 and rate > best_rate:
                best, best_rate = c, rate
        price_col = best

    if not ts_col or not price_col:
        raise ValueError(f"Δεν βρέθηκαν στήλες Timestamp/Price στο DAM CSV. Columns: {list(df.columns)}")
    return ts_col, price_col

def load_dam_quarterly_endtime(dam_csv_path: str, month: str):
    """
    Διαβάζει το Energy Charts CSV, βρίσκει header, θεωρεί ότι το timestamp είναι
    ΗΔΗ local START time ανά 15λεπτο (00:00, 00:15, ..., 23:45) και
    ΔΕΝ το μετακινεί -15'.

    Επιστρέφει: TIMESTAMP (local START), DAM Price (€/MWh), dup_idx
    """
    fn = "load_dam_prices_15min"
    try:
        header_line = _find_header_line(dam_csv_path)
        dam = pd.read_csv(dam_csv_path, sep=None, engine="python", encoding="utf-8-sig", header=header_line)
        dam = dam.loc[:, ~dam.columns.astype(str).str.fullmatch(r"Unnamed: \d+")]
        dam.columns = [str(c).strip() for c in dam.columns]

        ts_col, price_col = _infer_dam_columns(dam)

        # Parse as UTC-if-possible, then σε Europe/Athens, ΧΩΡΙΣ -15 λεπτά
        ts_aware = pd.to_datetime(dam[ts_col], errors="coerce", utc=True)
        if ts_aware.isna().all():
            start_local = pd.to_datetime(dam[ts_col], errors="coerce")  # naive local
        else:
            start_local = ts_aware.dt.tz_convert("Europe/Athens").dt.tz_localize(None)

        price = pd.to_numeric(
            dam[price_col].astype(str).str.replace(",", ".", regex=False),
            errors="coerce"
        )

        out = pd.DataFrame({
            "TIMESTAMP": start_local,
            "DAM Price (€/MWh)": price
        }).dropna(subset=["TIMESTAMP"])

        # keep rows from 2025-10-01 and month filter
        lb = pd.Timestamp("2025-10-01 00:00")
        out = out[out["TIMESTAMP"] >= lb]
        out = out[out["TIMESTAMP"].dt.strftime("%Y-%m") == month].copy()

        # DO NOT sort: keep CSV order; dup_idx for duplicate times (DST fallback)
        out["dup_idx"] = out.groupby("TIMESTAMP").cumcount()

        log(fn, f"DAM 15' rows after filters: {len(out)} for {month}")
        return out.reset_index(drop=True)

    except Exception as e:
        log(fn, f"ERROR {e}")
        return None

# ===== Production =====

def read_production_data(file_path):
    fn = "read_production_data"
    try:
        df = pd.read_csv(file_path, sep=None, engine="python", encoding="utf-8-sig")
        log(fn, f"read {file_path}: {len(df)} rows")
        return df
    except Exception as e:
        log(fn, f"ERROR {file_path}: {e}")
        return None

def filter_monthly_data(df, month):
    """
    Βασικό filter: κρατάμε μόνο γραμμές για τον μήνα 'YYYY-MM'.
    Έχεις και ειδικούς κανόνες για 01/10 και DST, αλλά η κύρια λογική pairing
    γίνεται στη calculate_daily_summary_quarterly.
    """
    fn = "filter_monthly_data"

    if 'TIMESTAMP' not in df.columns or 'Μήνας' not in df.columns:
        log(fn, "missing TIMESTAMP/Μήνας")
        return pd.DataFrame()

    df = df[df['Μήνας'] == month].copy()
    if df.empty:
        log(fn, f"filtered 0 for {month} (by Μήνας)")
        return df

    ts = pd.to_datetime(
        df['TIMESTAMP'],
        format="%d/%m/%Y %H:%M",
        errors='coerce',
        dayfirst=True
    )
    df = df[ts.notna()].copy()
    df['TIMESTAMP'] = ts[ts.notna()]

    df = df[df['TIMESTAMP'].dt.minute.isin([0, 15, 30, 45])].copy()

    # SPECIAL CASE: 2025-10-01, πετάμε START 00:15/00:30/00:45
    mask_0110 = (
        (df['TIMESTAMP'].dt.date == pd.Timestamp("2025-10-01").date()) &
        (df['TIMESTAMP'].dt.hour == 0) &
        (df['TIMESTAMP'].dt.minute.isin([15, 30, 45]))
    )
    if mask_0110.any():
        df = df[~mask_0110].copy()

    df = df.sort_index(kind='stable')
    dup_mask = df.duplicated(subset=['TIMESTAMP'], keep='first')
    if dup_mask.any():
        df = df[~dup_mask].copy()

    log(fn, f"filtered {len(df)} for month={month} (START window + special rules)")
    return df

# (align_with_dam_quarterly μένει αλλά δεν χρησιμοποιείται πλέον στην κύρια ροή,
# το pairing γίνεται μέσα στη calculate_daily_summary_quarterly)

def align_with_dam_quarterly(df_prod, dam_df_15m):
    fn = "align_with_dam_quarterly_new"
    try:
        df = df_prod.copy()
        df['TIMESTAMP'] = pd.to_datetime(
            df['TIMESTAMP'],
            format="%d/%m/%Y %H:%M",
            errors='coerce',
            dayfirst=True
        )
        df = df.dropna(subset=['TIMESTAMP'])

        if 'ΕΝΕΡΓΕΙΑ (kWh)' in df.columns:
            s = df['ΕΝΕΡΓΕΙΑ (kWh)'].astype(str).str.strip()
            both = s.str.contains(r'\.') & s.str.contains(r',')
            s_both = s[both].str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
            s_else = s[~both].str.replace(',', '.', regex=False)
            s_clean = pd.concat([s_both, s_else]).reindex(s.index)
            df['ΕΝΕΡΓΕΙΑ (kWh)'] = pd.to_numeric(s_clean, errors='coerce').fillna(0.0)

        df = df.sort_values('TIMESTAMP', kind='mergesort')
        df = df.drop_duplicates(subset=['TIMESTAMP'], keep='first')

        dam = dam_df_15m.copy()
        dam = dam.dropna(subset=['TIMESTAMP'])
        dam = dam.sort_values('TIMESTAMP', kind='mergesort')
        dam = dam.drop_duplicates(subset=['TIMESTAMP'], keep='first')
        dam['DAM Price (€/MWh)'] = pd.to_numeric(dam['DAM Price (€/MWh)'], errors='coerce')

        rows = []
        all_days = sorted(df['TIMESTAMP'].dt.date.unique())

        for D in all_days:
            day_ts = pd.Timestamp(D)
            next_day_ts = day_ts + pd.Timedelta(days=1)

            is_oct_first_2025 = (day_ts == pd.Timestamp("2025-10-01"))

            if is_oct_first_2025:
                prod_main = df[
                    (df['TIMESTAMP'] >= day_ts + pd.Timedelta(hours=1, minutes=15)) &
                    (df['TIMESTAMP'] <= day_ts + pd.Timedelta(hours=23, minutes=45))
                ]
            else:
                prod_main = df[
                    (df['TIMESTAMP'] >= day_ts + pd.Timedelta(minutes=15)) &
                    (df['TIMESTAMP'] <= day_ts + pd.Timedelta(hours=23, minutes=45))
                ]

            prod_next_midnight = df[df['TIMESTAMP'] == next_day_ts]
            prod_day = pd.concat([prod_main, prod_next_midnight]).sort_values('TIMESTAMP')

            if prod_day.empty:
                continue

            if is_oct_first_2025:
                dam_day = dam[
                    (dam['TIMESTAMP'] >= day_ts + pd.Timedelta(hours=1)) &
                    (dam['TIMESTAMP'] <= day_ts + pd.Timedelta(hours=23, minutes=45))
                ]
            else:
                dam_day = dam[
                    (dam['TIMESTAMP'] >= day_ts) &
                    (dam['TIMESTAMP'] <= day_ts + pd.Timedelta(hours=23, minutes=45))
                ]

            if dam_day.empty:
                log(fn, f"⚠️ Καμία DAM τιμή για ημέρα {D}")
                continue

            prod_day = prod_day.reset_index(drop=True)
            dam_day  = dam_day.reset_index(drop=True)

            n_pairs = min(len(prod_day), len(dam_day))
            if n_pairs == 0:
                continue
            if len(prod_day) != len(dam_day):
                log(fn, f"⚠️ Μη ίσο πλήθος για {D}: prod={len(prod_day)}, dam={len(dam_day)} -> χρησιμοποιώ {n_pairs}")

            prod_day = prod_day.iloc[:n_pairs]
            dam_day  = dam_day.iloc[:n_pairs]

            day_df = pd.DataFrame({
                "Ημέρα": [day_ts.date()] * n_pairs,
                "TIMESTAMP_PROD": prod_day["TIMESTAMP"].values,
                "TIMESTAMP_DAM": dam_day["TIMESTAMP"].values,
                "ΕΝΕΡΓΕΙΑ (kWh)": prod_day["ΕΝΕΡΓΕΙΑ (kWh)"].values,
                "DAM Price (€/MWh)": dam_day["DAM Price (€/MWh)"].values,
            })

            rows.append(day_df)

        if not rows:
            log(fn, "⚠️ Δεν προέκυψαν ζευγάρια παραγωγής–DAM")
            return pd.DataFrame(columns=["Ημέρα","TIMESTAMP_PROD","TIMESTAMP_DAM","ΕΝΕΡΓΕΙΑ (kWh)","DAM Price (€/MWh)"])

        result = pd.concat(rows, ignore_index=True)
        log(fn, f"OK align_with_dam_quarterly_new: {len(result)} γραμμές ζευγαριών")
        return result

    except Exception as e:
        log(fn, f"ERROR {e}")
        return None

def calculate_daily_summary_quarterly(df_prod, df_dam_15m, producer_row, month):
    """
    ΝΕΑ ΛΟΓΙΚΗ – pairing ανά index:
      - Prod(END): D 00:15..23:45 + (D+1) 00:00
      - DAM(START): D 00:00..23:45
      - P[i] ↔ DAM[i] χωρίς shift.

    1/10: πετάμε τις 00:15/00:30/00:45
    26/10: κρατάμε μόνο την πρώτη εμφάνιση στα διπλά 03:00–04:00 (prod & DAM).
    """
    fn = "calculate_daily_summary_15m_by_index"
    try:
        month_str = month

        # =============== 1. ΠΑΡΑΓΩΓΗ (END TS) ================
        prod = df_prod.copy()
        prod['END_TS'] = pd.to_datetime(
            prod['TIMESTAMP'],
            format="%d/%m/%Y %H:%M",
            errors='coerce',
            dayfirst=True
        )
        prod = prod.dropna(subset=['END_TS'])
        prod = prod[prod['END_TS'].dt.strftime("%Y-%m") == month_str].copy()
        if prod.empty:
            log(fn, f"no production rows for {month_str}")
            return None, None

        prod['ΕΝΕΡΓΕΙΑ (kWh)'] = pd.to_numeric(
            prod['ΕΝΕΡΓΕΙΑ (kWh)'], errors='coerce'
        ).fillna(0.0)
        prod = prod.sort_values('END_TS').reset_index(drop=True)

        # =============== 2. DAM (START TS) ================
        dam = df_dam_15m.copy()
        dam['START_TS'] = pd.to_datetime(dam['TIMESTAMP'], errors='coerce')
        dam = dam.dropna(subset=['START_TS'])
        dam = dam[dam['START_TS'].dt.strftime("%Y-%m") == month_str].copy()
        if dam.empty:
            log(fn, f"no DAM rows for {month_str}")
            return None, None

        dam = dam.sort_values('START_TS').reset_index(drop=True)

        # =============== 3. Λίστα ημερών ================
        prod['day'] = prod['END_TS'].dt.date
        days = sorted({d for d in prod['day'] if str(d).startswith(month_str)})
        if not days:
            log(fn, f"no days in production for {month_str}")
            return None, None

        all_quarters = []

        for D in days:
            D_ts = pd.Timestamp(str(D))
            D_next = D_ts + pd.Timedelta(days=1)

            # ---- Prod: (D,00:00]..(D+1,00:00] => 00:15..23:45 + next 00:00 ----
            day_prod = prod[
                (prod['END_TS'] > D_ts) & (prod['END_TS'] <= D_next)
            ].copy()

            # 01/10/2025: πετάμε END 00:15/00:30/00:45
            if D == pd.Timestamp("2025-10-01").date():
                mask_skip = (
                    (day_prod['END_TS'].dt.date == D) &
                    (day_prod['END_TS'].dt.hour == 0) &
                    (day_prod['END_TS'].dt.minute.isin([15, 30, 45]))
                )
                day_prod = day_prod[~mask_skip].copy()

            # 26/10/2025: intervals 03:00–04:00 → END 03:15,03:30,03:45,04:00 (κρατάμε πρώτη)
            if D == pd.Timestamp("2025-10-26").date():
                mask_win = (
                    ((day_prod['END_TS'].dt.hour == 3) & day_prod['END_TS'].dt.minute.isin([15, 30, 45])) |
                    ((day_prod['END_TS'].dt.hour == 4) & (day_prod['END_TS'].dt.minute == 0))
                )
                dup = day_prod[mask_win].duplicated(subset=['END_TS'], keep='first')
                day_prod = day_prod.drop(index=day_prod[mask_win].loc[dup].index)

            day_prod = day_prod.sort_values('END_TS').reset_index(drop=True)
            if day_prod.empty:
                continue

            # ---- DAM: D 00:00..23:45 ----
            day_dam = dam[
                (dam['START_TS'] >= D_ts) &
                (dam['START_TS'] <= D_ts + pd.Timedelta(hours=23, minutes=45))
            ].copy()

            # 26/10/2025 DAM: διπλά 03:00–03:45 → κρατάμε πρώτη εμφάνιση
            if D == pd.Timestamp("2025-10-26").date():
                mask_dam_win = (
                    (day_dam['START_TS'].dt.hour == 3) &
                    (day_dam['START_TS'].dt.minute.isin([0, 15, 30, 45]))
                )
                dup_dam = day_dam[mask_dam_win].duplicated(subset=['START_TS'], keep='first')
                day_dam = day_dam.drop(index=day_dam[mask_dam_win].loc[dup_dam].index)

            day_dam = day_dam.sort_values('START_TS').reset_index(drop=True)
            if day_dam.empty:
                continue

            # ---- P[i] ↔ DAM[i] ----
            n_p = len(day_prod)
            n_d = len(day_dam)
            n = min(n_p, n_d)
            if n == 0:
                continue
            if n_p != n_d:
                log(fn, f"Length mismatch {D}: prod={n_p}, dam={n_d}, using first {n}")

            day_prod = day_prod.iloc[:n].copy()
            day_dam = day_dam.iloc[:n].copy()

            kwh = day_prod['ΕΝΕΡΓΕΙΑ (kWh)'].to_numpy()
            price = day_dam['DAM Price (€/MWh)'].to_numpy()
            value_eur = (kwh * price) / 1000.0

            per_quarter = pd.DataFrame({
                'Περίοδος εκκαθάρισης': [D] * n,
                'ΕΝΕΡΓΕΙΑ (kWh)': kwh,
                'Αξία ενέργειας βάσει μετρήσεων (€)': value_eur
            })
            all_quarters.append(per_quarter)

        if not all_quarters:
            log(fn, f"no quarter rows after pairing for {month_str}")
            return None, None

        df_all = pd.concat(all_quarters, ignore_index=True)

        # =============== 4. ΗΜΕΡΗΣΙΑ ΣΥΝΟΛΑ ================
        df_daily = df_all.groupby('Περίοδος εκκαθάρισης', as_index=False).agg({
            'ΕΝΕΡΓΕΙΑ (kWh)': 'sum',
            'Αξία ενέργειας βάσει μετρήσεων (€)': 'sum'
        })

        df_daily['Περίοδος εκκαθάρισης'] = pd.to_datetime(
            df_daily['Περίοδος εκκαθάρισης']
        ).dt.strftime('%d/%m/%y')

        rate = float(producer_row['Μοναδιαία Χρέωση ΦοΣΕ'].values[0])
        df_daily['Προμήθεια GREEN VALUE (€)'] = (
            df_daily['ΕΝΕΡΓΕΙΑ (kWh)'] / 1000.0 * rate
        ).round(2)

        df_daily['Μεσοσταθμική Τιμή Αγοράς κατά τις ώρες παραγωγής του σταθμού'] = df_daily.apply(
            lambda row: 0 if row['ΕΝΕΡΓΕΙΑ (kWh)'] == 0 else round(
                (row['Αξία ενέργειας βάσει μετρήσεων (€)'] / row['ΕΝΕΡΓΕΙΑ (kWh)']) * 1000.0, 2
            ),
            axis=1
        )

        sum_energy = round(float(df_daily['ΕΝΕΡΓΕΙΑ (kWh)'].sum()), 2)
        sum_value  = round(float(df_daily['Αξία ενέργειας βάσει μετρήσεων (€)'].sum()), 2)
        sum_prov   = round(float(df_daily['Προμήθεια GREEN VALUE (€)'].sum()), 2)

        summary_row = pd.DataFrame([{
            'Περίοδος εκκαθάρισης': 'Σύνολο',
            'ΕΝΕΡΓΕΙΑ (kWh)': sum_energy,
            'Αξία ενέργειας βάσει μετρήσεων (€)': sum_value,
            'Προμήθεια GREEN VALUE (€)': sum_prov,
            'Μεσοσταθμική Τιμή Αγοράς κατά τις ώρες παραγωγής του σταθμού': (
                round((sum_value / sum_energy) * 1000.0, 2) if sum_energy > 0 else 0
            )
        }])

        df_final = pd.concat([df_daily, summary_row], ignore_index=True)
        log(fn, f"daily rows={len(df_final)} for month={month_str}")
        return df_final, (sum_energy, sum_value, sum_prov)

    except Exception as e:
        log(fn, f"ERROR {e}")
        return None, None

# ===== Output folders =====
def make_base_dirs(month):
    root = os.path.join('ΕΝΗΜΕΡΩΤΙΚΑ_ΣΗΜΕΙΩΜΑΤΑ', month)
    xlsx_dir = os.path.join(root, "XLSX")
    pdf_dir  = os.path.join(root, "PDF")
    os.makedirs(xlsx_dir, exist_ok=True)
    os.makedirs(pdf_dir,  exist_ok=True)
    return root, xlsx_dir, pdf_dir

def determine_pdf_subfolder_name(email, email_to_companies, email_to_customs):
    customs   = email_to_customs.get(email, [])
    companies = email_to_companies.get(email, [])
    return clipped_folder_name(customs, companies, limit=MAX_FOLDER_CHARS)

def xlsx_filename(company_name, month):
    return clipped_filename(company_name, month, "xlsx", limit=MAX_FILENAME_CHARS)

def pdf_filename(company_name, month):
    return clipped_filename(company_name, month, "pdf",  limit=MAX_FILENAME_CHARS)

# ===== Excel generation from template =====
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

def generate_invoice_excel(df_daily_energy, summary, producer_row, month, xlsx_output_dir):
    fn = "generate_invoice_excel"
    try:
        year, month_number = month.split('-')
        greek_months_genitive = {
            '01': 'Ιανουαρίου','02':'Φεβρουαρίου','03':'Μαρτίου','04':'Απριλίου',
            '05': 'Μαΐου','06':'Ιουνίου','07':'Ιουλίου','08':'Αυγούστου',
            '09': 'Σεπτεμβρίου','10':'Οκτωβρίου','11':'Νοεμβρίου','12':'Δεκεμβρίου'
        }

        template_path = 'Invoice_GREEN_VALUE_01.xlsx'
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Λείπει template: {template_path}")
        
        company_name = str(producer_row['Εταιρεία'].values[0])
        email_value  = str(producer_row['Email'].values[0]) if 'Email' in producer_row else ''
        iban = producer_row['IBAN'].values[0] if 'IBAN' in producer_row else ''
        rate = float(producer_row['Μοναδιαία Χρέωση ΦοΣΕ'].values[0])
        sum_energy, sum_value, sum_prov = summary

        out_name = xlsx_filename(company_name, month)
        xlsx_path = os.path.join(xlsx_output_dir, out_name)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        wb = load_workbook(template_path)
        ws = wb.active

        ws.row_dimensions[1].height = 160
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 40
        ws.row_dimensions[5].height = 50  
    
        for rng in ('C2:G2',):
            try: ws.unmerge_cells(rng)
            except Exception: pass

        for col in range(1, 9):
            c = ws.cell(row=2, column=col)
            if not isinstance(c, MergedCell):
                c.value = None

        ws.merge_cells('C2:G2')
        ws['C2'] = f'Ενημερωτικό Σημείωμα {greek_months_genitive.get(month_number, "")} {year}'
        ws['C2'].font = xls.Font(name="Century Gothic", bold=True, underline='single', size=20)
        ws['C2'].alignment = xls.Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.merge_cells('G1:H1')
        ws['G1'] = (
            'Φορέας Σωρευτικής Εκπροσώπησης ΑΠΕ (Φο.Σ.Ε.)\n'
            'Διεύθυνση: Φιλοπάππου 19, Αθήνα 11741, Ελλάδα \n'
            'ΑΦΜ: 801961185, ΓΕΜΗ: 167104201000, ΔΟΥ:ΦΑΕ Αθηνών \n'
            'e-mail: info@greenvalue.gr'
        )
        ws['G1'].font = xls.Font(name="Century Gothic", size=14)
        ws['G1'].alignment = xls.Alignment(wrap_text=True, vertical='top')

        ws.row_dimensions[7].height = 14

        needed = ['Α.Μ. ΑΠΕ','Εταιρεία','ΑΦΜ','ΔΟΥ','Διεύθυνση','Email','Τεχνολογία']
        vals = producer_row.iloc[0][needed].tolist()
        for cell_ref, val in zip(['B4','C4','D4','E4','F4','G4','H4'], vals):
            ws[cell_ref] = val
        ws['C4'].font = xls.Font(name="Century Gothic", size=13)

        today = pd.Timestamp.today().strftime('%d/%m/%y')
        start_date = pd.to_datetime(df_daily_energy.iloc[0]['Περίοδος εκκαθάρισης'], dayfirst=True)
        end_date   = pd.to_datetime(df_daily_energy.iloc[-2]['Περίοδος εκκαθάρισης'], dayfirst=True)
        ws['B6'] = today
        ws['C6'] = 'Αρχική'
        ws.merge_cells('D6:F6')
        ws['D6'] = f"{start_date.strftime('%d/%m/%y')}-{end_date.strftime('%d/%m/%y')}"

        for row in ws.iter_rows(min_row=10, max_row=41, min_col=3, max_col=7):
            for cell in row:
                cell.value = None

        start_row = 10
        for r_idx, row_vals in enumerate(df_daily_energy.values, start=start_row):
            for c_idx, value in enumerate(row_vals, start=3):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

        first_data_row = 10
        last_data_row  = first_data_row + len(df_daily_energy) - 1
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(row=r, column=7).alignment = xls.Alignment(horizontal='center', vertical='center')

        ws['H45'] = round(sum_energy / 1000, 2); ws['H45'].number_format = '#,##0.00'
        ws['H46'] = round(sum_value, 2);         ws['H46'].number_format = '#,##0.00'
        ws['C52'] = iban; ws['C52'].font = xls.Font(name="Century Gothic", size=14, bold=True)
        ws['C53'] = (pd.Timestamp.today() + pd.Timedelta(days=2)).strftime('%d/%m/%y')
        ws['D45'] = rate
        ws['D46'] = round(sum_prov, 2); ws['D46'].number_format = '#,##0.00'

        try:
            ws.row_dimensions[58].height = 10
            _add_daily_plot(ws, df_daily_energy, anchor_cell="B59", color_hex="#22A052")
        except Exception as e:
            log(fn, f"Plot warning: {e}")

        fix_header_rows(ws)
        wb.save(xlsx_path)
        log(fn, f"XLSX: {xlsx_path}")
        return xlsx_path, company_name, email_value

    except Exception as e:
        log(fn, f"ERROR {e}")
        print(f"❌ Excel generation failed: {e}")
        return None, None, None

# ===== PDF export: Excel first (xlwings), then LibreOffice fallback =====
# def _find_soffice_path() -> str | None:
#     p = shutil.which("soffice")
#     if p:
#         return p
#     default = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
#     if os.path.exists(default):
#         return default
#     return None

def _verify_pdf(path: str, min_bytes: int = 500):
    return os.path.exists(path) and os.path.getsize(path) >= min_bytes

def export_to_pdf_with_excel(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    fn = "export_to_pdf_excel"
    try:
        import xlwings as xw
    except Exception as e:
        log(fn, f"xlwings import failed: {e}")
        return False, "excel-not-available"

    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(os.path.abspath(xlsx_path))
        sht = wb.sheets.active
        try:
            sht.api.PageSetup.Zoom = False
            sht.api.PageSetup.FitToPagesWide = 1
        except Exception:
            pass

        out_pdf = os.path.abspath(pdf_path)
        Path(os.path.dirname(out_pdf)).mkdir(parents=True, exist_ok=True)

        # xlwings wb.to_pdf (if available)
        try:
            if hasattr(wb, "to_pdf"):
                wb.to_pdf(out_pdf)
                export_failed = False
            else:
                raise AttributeError("wb.to_pdf not available")
        except Exception as e_to_pdf:
                log(fn, f"wb.to_pdf failed: {e_to_pdf}; trying AppleScript fallback...")

        # Close workbook
        try:
            wb.close()
        except Exception:
            pass

        # Check if PDF was created
        ok = _verify_pdf(out_pdf)
        if ok:
            log(fn, f"OK via Excel -> {out_pdf}")
            return True, "excel"
        else:
            log(fn, f"Excel export produced no/empty file at: {out_pdf}")
            return False, "excel-empty"
    except Exception as e:
        log(fn, f"ERROR (Excel outer): {e}")
        return False, f"excel-error:{e}"
    finally:
        try:
            app.quit()
        except Exception:
            pass

# def export_to_pdf_with_libreoffice(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    """
    Export XLSX -> PDF via LibreOffice (απλό convert-to pdf, χωρίς extra JSON options),
    για να μην κολλάει η soffice και να αποφεύγεται το timeout.
    """
    fn = "export_to_pdf_libreoffice"
    soffice = _find_soffice_path()
    if not soffice:
        log(fn, "LibreOffice 'soffice' not found.")
        return False, "lo-missing"

    outdir = os.path.abspath(os.path.dirname(pdf_path))
    Path(outdir).mkdir(parents=True, exist_ok=True)

    # ΑΠΛΟ convert-to pdf (χωρίς filter options που έσπαγαν την εντολή)
    cmd = [
        soffice, "--headless", "--norestore", "--nolockcheck",
        "--convert-to", "pdf",
        "--outdir", outdir,
        os.path.abspath(xlsx_path)
    ]

    try:
        # δίνουμε λίγο μεγαλύτερο περιθώριο, π.χ. 300s
        res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=300)
        if res.returncode != 0:
            log(fn, f"ERROR rc={res.returncode}\nstdout:\n{res.stdout}\nstderr:\n{res.stderr}")
            return False, f"lo-error-rc{res.returncode}"

        produced = os.path.join(outdir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")

        # Αν το τελικό όνομα είναι διαφορετικό, κάνουμε rename
        if os.path.abspath(produced) != os.path.abspath(pdf_path):
            try:
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                os.replace(produced, pdf_path)
            except Exception as e:
                log(fn, f"Rename error: {e}")
                return False, f"lo-rename-error:{e}"

        time.sleep(0.2)
        ok = _verify_pdf(pdf_path) if '_verify_pdf' in globals() else (os.path.getsize(pdf_path) > 500)
        if ok:
            log(fn, f"OK via LibreOffice -> {pdf_path}")
            return True, "libreoffice"
        else:
            log(fn, f"LibreOffice produced no/empty file at: {pdf_path}")
            return False, "lo-empty"
    except subprocess.TimeoutExpired as e:
        log(fn, f"TIMEOUT (LibreOffice): {e}")
        return False, "lo-timeout"
    except Exception as e:
        log(fn, f"ERROR (LibreOffice): {e}")
        return False, f"lo-exception:{e}"

def export_to_pdf(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    ok, how = export_to_pdf_with_excel(xlsx_path, pdf_path)
    if ok:
        return True, how
    # LibreOffice export is commented out, so return the Excel failure
    return False, how

# ===== Main flow =====
def timologia(month):
    producers_df = load_producers('producers.xlsx')
    if producers_df is None:
        print("Αποτυχία: producers.xlsx"); return

    email_to_companies, email_to_customs = build_email_groups(producers_df)

    dam_file = 'energy-charts_Electricity_production_and_spot_prices_in_Greece_in_2025.csv'
    df_dam_15m = load_dam_quarterly_endtime(dam_file, month)
    if df_dam_15m is None or df_dam_15m.empty:
        print("Αποτυχία: DAM 15' prices"); return

    root, xlsx_dir, pdf_dir = make_base_dirs(month)

    base_folder = 'ΠΑΡΑΓΩΓΗ'
    if not os.path.isdir(base_folder):
        print(f"Λείπει φάκελος: {base_folder}"); return

    for filename in os.listdir(base_folder):
        if not (filename.startswith('ΠΑΡΑΓΩΓΗ_') and filename.endswith('.csv')):
            continue

        print(f"\n=== Επεξεργασία αρχείου: {filename} ===")

        file_path = os.path.join(base_folder, filename)
        m = re.match(r'ΠΑΡΑΓΩΓΗ_(.+)\.csv', filename)
        if not m:
            log("timologia", f"Bad filename: {filename}")
            print("  -> SKIP: Bad filename pattern")
            continue

        company_key = m.group(1)
        prod_row = producers_df[producers_df['normalized_name'] == normalize_name(company_key)]
        if prod_row.empty:
            log("timologia", f"No producer for {filename}")
            print("  -> SKIP: Δεν βρέθηκε παραγωγός στο producers.xlsx για αυτό το filename")
            continue

        company_name = str(prod_row['Εταιρεία'].values[0])
        print(f"  Εταιρεία: {company_name}")

        df = read_production_data(file_path)
        if df is None:
            print("  -> SKIP: read_production_data επέστρεψε None")
            continue

        df = filter_monthly_data(df, month)
        if df.empty:
            print(f"  -> SKIP: Δεν υπάρχουν γραμμές παραγωγής για μήνα {month} (στήλη 'Μήνας')")
            continue

        df_daily, summary = calculate_daily_summary_quarterly(df, df_dam_15m, prod_row, month)
        if df_daily is None:
            print("  -> SKIP: calculate_daily_summary_quarterly επέστρεψε None (π.χ. δεν βρέθηκαν ζευγάρια P×DAM)")
            continue

        xlsx_path, company_name, email_value = generate_invoice_excel(df_daily, summary, prod_row, month, xlsx_dir)
        if not xlsx_path:
            print("  -> SKIP: generate_invoice_excel απέτυχε (δες το μήνυμα '❌ Excel generation failed: ...')")
            continue

        email_key = (email_value or "NO_EMAIL").strip() or "NO_EMAIL"
        subfolder = determine_pdf_subfolder_name(email_key, email_to_companies, email_to_customs)
        target_dir = os.path.join(pdf_dir, subfolder[:MAX_FOLDER_CHARS])
        os.makedirs(target_dir, exist_ok=True)

        pdf_name = pdf_filename(company_name, month)
        pdf_path = os.path.join(target_dir, pdf_name)

        ok, how = export_to_pdf(xlsx_path, pdf_path)
        status = "✅ PDF" if ok else "❌ PDF"
        print(f"  {status} [{how}] → {pdf_path}")
        if not ok:
            log("timologia", f"PDF export failed for: {xlsx_path} | method={how}")
    
    print(f"\nΈτοιμο. Δες: ΕΝΗΜΕΡΩΤΙΚΑ_ΣΗΜΕΙΩΜΑΤΑ/{month}/XLSX και /PDF")

if __name__ == '__main__':
    month_input = input("Δώσε μήνα (YYYY-MM): ").strip()
    if not re.match(r'^\d{4}-(0[1-9]|1[0-2])$', month_input):
        print("Μη έγκυρη μορφή. Παράδειγμα: 2025-10")
    else:
        timologia(month_input)