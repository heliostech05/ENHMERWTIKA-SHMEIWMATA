#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inspect merged cells and structure in WEEKLY_Invoice_GREEN_VALUE_01.xlsx
"""

from pathlib import Path
from openpyxl import load_workbook

TEMPLATE_FILE = Path(__file__).parent / "WEEKLY_Invoice_GREEN_VALUE_01.xlsx"

if not TEMPLATE_FILE.exists():
    print(f"❌ Template not found: {TEMPLATE_FILE}")
    exit(1)

wb = load_workbook(TEMPLATE_FILE)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"\n=== Merged Cells ===")
for rg in ws.merged_cells.ranges:
    print(f"  {rg.coord}")

print(f"\n=== Key Cells to Update ===")
key_cells = ['C1', 'G2', 'B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'D5', 'H42', 'H43', 'C48', 'C49', 'D42', 'D43']
for cell_ref in key_cells:
    cell = ws[cell_ref]
    is_merged = isinstance(cell, type(cell)) and hasattr(cell, 'is_merged') and cell.is_merged
    print(f"  {cell_ref}: value={cell.value}, merged={is_merged}")

print(f"\n=== Data Area (C8:G39) ===")
for rg in ws.merged_cells.ranges:
    bounds = rg.bounds
    if bounds[0] >= 3 and bounds[0] <= 7 and bounds[1] >= 8 and bounds[1] <= 39:
        print(f"  Merged range in data area: {rg.coord}")

print("\n✅ Inspection complete")
