#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
enimerwtika_weekly_sithya.py

Εβδομαδιαία ενημερωτικά ΜΟΝΟ για ΣΗΘΥΑ.

ΛΟΓΙΚΗ
------
- Χρησιμοποιεί ΑΚΡΙΒΩΣ τη λογική υπολογισμού του μηνιαίου script
  (pairing 15λέπτων παραγωγής–DAM, DST, πράξεις κ.λπ.) μέσω της
  συνάρτησης calculate_daily_summary_quarterly, η οποία εδώ
  έχει αντιγραφεί από το μηνιαίο.

- Για κάθε παραγωγό ΣΗΘΥΑ:
    1) Τρέχουμε τη μηνιαία calculate_daily_summary_quarterly
       για ολόκληρο τον μήνα.
    2) Φιλτράρουμε τις ημερομηνίες μόνο στο διάστημα
       [start_date .. end_date].
    3) Ξαναϋπολογίζουμε τα σύνολα μόνο για αυτές τις μέρες.
    4) Φτιάχνουμε Excel/PDF με εβδομαδιαίο τίτλο.

- ΠΡΟΣΟΧΗ: Η εβδομάδα πρέπει να είναι μέσα στον ίδιο μήνα.

Απαιτεί:
- ADMIE_MERGED_TIMOLOGIA.py (το μηνιαίο) ΔΕΝ χρειάζεται να γίνει import.
  Εδώ έχουν αντιγραφεί οι απαραίτητες συναρτήσεις (DAM, παραγωγή, calculation).
"""

import os
import re
import shutil
import subprocess
import time
from pathlib import Path
from collections import defaultdict
import unicodedata
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

# Προαιρετικά για PDF μέσω Excel
try:
    import xlwings as xw  # noqa
    _HAS_XLWINGS = True
except Exception:
    _HAS_XLWINGS = False

# =================== Paths / Globals ===================

BASE_DIR = Path(__file__).resolve().parent

TEMPLATE_FILE   = BASE_DIR / "WEEKLY_Invoice_GREEN_VALUE_01.xlsx"
PRODUCERS_XLSX  = BASE_DIR / "producers.xlsx"
DAM_FILE_2025   = BASE_DIR / "energy-charts_Electricity_production_and_spot_prices_in_Greece_in_2026.csv"

PROD_DIR        = BASE_DIR / "ΠΑΡΑΓΩΓΗ"
DOWNLOADS_DIR   = BASE_DIR / "downloads"   # downloads/<YYYY-MM>/GREEN_VE6*.csv

LOG_BASE        = BASE_DIR / "logs" / "timologia_weekly"
LOG_BASE.mkdir(parents=True, exist_ok=True)

MAX_FOLDER_CHARS   = 120
MAX_FILENAME_CHARS = 140

WIN_RESERVED = {
    "CON","PRN","AUX","NUL",
    "COM1","COM2","COM3","COM4","COM5","COM6","COM7","COM8","COM9",
    "LPT1","LPT2","LPT3","LPT4","LPT5","LPT6","LPT7","LPT8","LPT9"
}


def log(name, msg):
    with open(LOG_BASE / f"{name}.txt", "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")


# =================== Helpers (names/paths) ===================

def sanitize_name(name: str) -> str:
    if name is None:
        return "UNTITLED"
    s = str(name)
    s = re.sub(r'[\\/*?:"<>|]', "", s)
    s = s.strip().rstrip(".")
    if not s:
        s = "UNTITLED"
    if s.upper() in WIN_RESERVED:
        s = "_" + s
    return s

def normalize_name(name: str) -> str:
    return re.sub(r'[\\._\\-\\s]', '', str(name).strip().lower())

def join_with_limit(parts, sep=" & ", limit=120):
    out, total = [], 0
    for p in parts:
        piece = p if not out else (sep + p)
        if total + len(piece) > limit:
            break
        out.append(p)
        total += len(piece)
    if out:
        return sep.join(out)
    return (parts[0] if parts else "UNTITLED")[:limit]

def clipped_folder_name(preferred_names, fallback_names, limit=MAX_FOLDER_CHARS):
    def build_name(items):
        uniq = sorted({sanitize_name(x) for x in items if x})
        return join_with_limit(uniq, sep=" & ", limit=limit)
    if preferred_names:
        return build_name(preferred_names)
    return build_name(fallback_names)

def clipped_filename_weekly(company_name: str, tag: str, ext: str, limit=MAX_FILENAME_CHARS):
    prefix = "ΕΒΔΟΜΑΔΙΑΙΟ_ΣΗΜΕΙΩΜΑ_"
    base = sanitize_name(company_name).replace(" ", "_")
    cand = f"{prefix}{base}_{tag}.{ext}"
    if len(cand) <= limit:
        return cand
    over = len(cand) - limit
    base_cut = base[:max(1, len(base) - over)]
    cand = f"{prefix}{base_cut}_{tag}.{ext}"
    if len(cand) > limit:
        trunk = f"{prefix}{tag}"
        cand = trunk[:limit - (len(ext) + 1)] + "." + ext
    return cand

def xlsx_filename_weekly(company_name, tag):
    return clipped_filename_weekly(company_name, tag, "xlsx", MAX_FILENAME_CHARS)

def pdf_filename_weekly(company_name, tag):
    return clipped_filename_weekly(company_name, tag, "pdf", MAX_FILENAME_CHARS)


# =================== Producers (SITHYA only) ===================

def load_producers_sithya(filepath=PRODUCERS_XLSX):
    fn = "load_producers_sithya"
    try:
        if not Path(filepath).exists():
            log(fn, f"❌ producers.xlsx not found: {filepath}")
            return None
        df = pd.read_excel(filepath, dtype={'Code': str})

        for col in ['Email','IBAN','Code']:
            if col not in df.columns:
                df[col] = ""
        needed = [
            'Εταιρεία','Email','Μοναδιαία Χρέωση ΦοΣΕ',
            'Α.Μ. ΑΠΕ','ΑΦΜ','ΔΟΥ','Διεύθυνση','Τεχνολογία','IBAN','Code'
        ]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            raise ValueError(f"Λείπουν στήλες: {missing}")

        tech = df['Τεχνολογία'].astype(str).str.strip().str.upper()
        aliases = {"ΣΗΘΥΑ","ΣΗΘ","ΣΗΘΥΑ/ΣΗΘ","ΣΗΘΥΑ - CHP","CHP","ΣΗΘ-YA"}
        mask = tech.isin(aliases) | tech.str.contains(r"\bΣΗΘΥΑ\b", regex=True)
        df = df[mask].copy()

        if df.empty:
            log(fn, "⚠️ Δεν βρέθηκαν παραγωγοί ΣΗΘΥΑ.")
            return df

        df['normalized_name'] = df['Εταιρεία'].astype(str).apply(normalize_name)
        # Also produce a per-producer log file so we have one document per company
        try:
            for _, prow in df.iterrows():
                comp_name = str(prow.get('Εταιρεία', '') or '').strip()
                if not comp_name:
                    continue
                safe = normalize_name(comp_name)
                try:
                    log(f"producer_{safe}", prow.to_dict())
                except Exception as e:
                    log(fn, f"failed to write individual log for {comp_name}: {e}")
        except Exception as e:
            log(fn, f"failed to produce per-producer logs: {e}")
        log(fn, f"OK ΣΗΘΥΑ producers: {len(df)}")
        return df
    except Exception as e:
        log(fn, f"ERROR {e}")
        return None

def build_email_groups(producers_df):
    email_to_companies = defaultdict(set)
    for _, row in producers_df.iterrows():
        email = (str(row.get('Email', '') or '').strip()) or "NO_EMAIL"
        comp  = str(row.get('Εταιρεία', '') or '').strip()
        if comp:
            email_to_companies[email].add(comp)
    email_to_companies = {em: sorted(v) for em, v in email_to_companies.items()}
    return email_to_companies, {}


# =================== GREEN_VE6 → ΠΑΡΑΓΩΓΗ ===================

def load_producers_basic(filepath=PRODUCERS_XLSX):
    fn = "load_producers_basic"
    try:
        if not Path(filepath).exists():
            log(fn, f"❌ producers.xlsx not found: {filepath}")
            return None
        df = pd.read_excel(filepath, dtype={'Code': str})
        if 'Code' not in df.columns or 'Εταιρεία' not in df.columns:
            raise ValueError("Το producers.xlsx πρέπει να έχει 'Code' και 'Εταιρεία'")
        df['Code']      = df['Code'].astype(str).str.strip()
        df['Εταιρεία']  = df['Εταιρεία'].astype(str).str.strip()
        return df[['Code','Εταιρεία']]
    except Exception as e:
        log(fn, f"ERROR {e}")
        return None

def get_latest_green_ve6_files(folder: Path):
    fn = "get_latest_green_ve6_files"
    date_to_files = defaultdict(list)
    csvs = [f for f in os.listdir(folder) if f.startswith("GREEN_VE6") and f.endswith(".csv")]
    log(fn, f"Βρέθηκαν {len(csvs)} αρχεία στο {folder}")
    for name in csvs:
        m = re.match(r"GREEN_VE6(\d{8})(\d)\.csv", name)
        if not m:
            continue
        date    = m.group(1)
        edition = int(m.group(2))
        date_to_files[date].append((edition, name))
    latest = []
    for date in sorted(date_to_files.keys()):
        files = sorted(date_to_files[date], reverse=True)
        latest.append(files[0][1])
        log(fn, f"{date} -> έκδοση {files[0][0]}: {files[0][1]}")
    return latest

def preprocess_timestamp_column(df):
    is_24 = df['TIMESTAMP'].astype(str).str.contains('24:00', regex=False)
    new_ts = df['TIMESTAMP'].astype(str)
    new_ts[is_24] = (
        pd.to_datetime(
            new_ts[is_24].str.replace('24:00','00:00'),
            format='%d/%m/%Y %H:%M',
            errors='coerce'
        ) + pd.Timedelta(days=1)
    ).dt.strftime('%d/%m/%Y %H:%M')
    df['TIMESTAMP'] = new_ts
    df['datetime']  = pd.to_datetime(df['TIMESTAMP'], format='%d/%m/%Y %H:%M', errors='coerce')
    return df

def safe_company_folder_name(name):
    return re.sub(r'[\\/*?:"<>|]', "", name.replace(" ", "_"))

def merge_with_existing_csv(group_df, out_file: Path):
    g = group_df.copy()
    g.set_index('TIMESTAMP', inplace=True)
    if out_file.exists():
        try:
            exist = pd.read_csv(out_file, delimiter=';', encoding='utf-8-sig')
            exist.set_index('TIMESTAMP', inplace=True)
            combined = exist[~exist.index.isin(g.index)]
            combined = pd.concat([combined, g])
        except Exception as e:
            log("merge_with_existing_csv", f"Σφάλμα ανάγνωσης {out_file}: {e}")
            combined = g
    else:
        combined = g
    combined = combined.reset_index()
    combined['datetime'] = pd.to_datetime(
        combined['TIMESTAMP'],
        format='%d/%m/%Y %H:%M',
        errors='coerce'
    )
    combined = combined.sort_values('datetime').drop(columns=['datetime'])
    return combined

def process_green_ve6_file(filepath: Path, producers_map: pd.DataFrame, output_folder: Path):
    fn = "process_green_ve6_file"
    log(fn, f"Επεξεργασία: {filepath.name}")
    try:
        df = pd.read_csv(filepath, delimiter=';', encoding='utf-8-sig', skiprows=1)
    except Exception as e:
        log(fn, f"Σφάλμα ανάγνωσης: {e}")
        return
    if 'ΚΩΔΙΚΟΣ ΕΔΡΕΘ' not in df.columns or 'TIMESTAMP' not in df.columns:
        log(fn, "Λείπει 'ΚΩΔΙΚΟΣ ΕΔΡΕΘ' ή 'TIMESTAMP'")
        return
    df = preprocess_timestamp_column(df)
    for code_value, group in df.groupby('ΚΩΔΙΚΟΣ ΕΔΡΕΘ'):
        code_str = str(code_value).strip()
        row = producers_map[producers_map['Code'] == code_str]
        if row.empty:
            log(fn, f"Άγνωστος Code={code_str} (δεν υπάρχει στο producers.xlsx)")
            continue
        company   = row['Εταιρεία'].values[0]
        safe_name = safe_company_folder_name(company)
        out_file  = output_folder / f"ΠΑΡΑΓΩΓΗ_{safe_name}.csv"
        final_df  = merge_with_existing_csv(group, out_file)
        final_df.to_csv(out_file, index=False, sep=';', encoding='utf-8-sig')
        log(fn, f"OK -> {out_file}")

def _month_iter(start_date: pd.Timestamp, end_date: pd.Timestamp):
    cur   = pd.Timestamp(start_date.year, start_date.month, 1)
    final = pd.Timestamp(end_date.year,   end_date.month,   1)
    while cur <= final:
        yield cur.year, cur.month
        cur = (cur + pd.offsets.MonthBegin(1))

def ensure_production_files(start_date: pd.Timestamp, end_date: pd.Timestamp):
    """
    Χτίζει/ενημερώνει ΠΑΡΑΓΩΓΗ_*.csv από downloads/<YYYY-MM>/GREEN_VE6*,
    για όλους τους μήνες που καλύπτουν το [start_date .. end_date].
    """
    fn = "ensure_production_files_multi"
    producers_map = load_producers_basic(PRODUCERS_XLSX)
    if producers_map is None or producers_map.empty:
        print("⚠️ Δεν μπόρεσα να διαβάσω producers.xlsx (ή λείπουν Code/Εταιρεία).")
        return

    PROD_DIR.mkdir(parents=True, exist_ok=True)
    total_sources = 0

    for y, m in _month_iter(start_date, end_date):
        month_tag    = f"{y}-{m:02d}"
        input_folder = DOWNLOADS_DIR / month_tag
        if not input_folder.is_dir():
            log(fn, f"skip: no downloads/{month_tag}")
            continue
        latest_files = get_latest_green_ve6_files(input_folder)
        if not latest_files:
            log(fn, f"skip: no GREEN_VE6 in {input_folder}")
            continue

        print(f"🔧 Χτίζω ΠΑΡΑΓΩΓΗ από {len(latest_files)} GREEN_VE6 αρχεία για {month_tag}...")
        for name in latest_files:
            process_green_ve6_file(input_folder / name, producers_map, PROD_DIR)
            total_sources += 1

    if total_sources:
        print("✅ Έτοιμο το ΠΑΡΑΓΩΓΗ/ (ενημερώθηκε για όλους τους μήνες)")
    else:
        print("ℹ️ Δεν βρέθηκαν νέα GREEN_VE6. Συνεχίζω με τα υπάρχοντα ΠΑΡΑΓΩΓΗ_*.csv.")


# =================== DAM utilities (αντιγραφή από μηνιαίο) ===================

HEADER_TS_KEYS    = ["date", "time", "timestamp", "cet", "ce(s)t", "gmt", "utc", "eet", "athens", "gmt+2"]
HEADER_PRICE_KEYS = ["price", "eur/mwh", "€/mwh", "auction", "day-ahead", "day ahead"]

def _find_header_line(path, max_scan=200):
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        for i in range(max_scan):
            line = f.readline()
            if not line:
                break
            low = line.strip().lower()
            if any(k in low for k in HEADER_TS_KEYS) and any(k in low for k in HEADER_PRICE_KEYS):
                return i
    return 1

def _infer_dam_columns(df: pd.DataFrame):
    cols = list(df.columns)
    lower = {c: c.lower() for c in cols}

    ts_cands = [c for c in cols if any(k in lower[c] for k in HEADER_TS_KEYS)]
    price_cands = [c for c in cols if any(k in lower[c] for k in HEADER_PRICE_KEYS)]

    ts_col = ts_cands[0] if ts_cands else None
    price_cands = [c for c in price_cands if c != ts_col]
    price_col = price_cands[0] if price_cands else None

    if not ts_col:
        best, best_rate = None, -1
        for c in cols:
            try:
                parsed = pd.to_datetime(df[c], errors="coerce")
                rate = parsed.notna().sum() / max(1, df[c].notna().sum())
                if rate >= 0.8 and rate > best_rate:
                    best, best_rate = c, rate
            except Exception:
                pass
        ts_col = best
    if not price_col:
        best, best_rate = None, -1
        for c in cols:
            if c == ts_col:
                continue
            s = pd.to_numeric(df[c].astype(str).str.replace(",", ".", regex=False), errors="coerce")
            rate = s.notna().sum() / max(1, df[c].notna().sum())
            if rate >= 0.8 and rate > best_rate:
                best, best_rate = c, rate
        price_col = best

    if not ts_col or not price_col:
        raise ValueError(f"Δεν βρέθηκαν στήλες Timestamp/Price στο DAM CSV. Columns: {list(df.columns)}")
    return ts_col, price_col

def load_dam_quarterly_endtime(dam_csv_path: str, month: str):
    """
    Διαβάζει το Energy Charts CSV, βρίσκει header, θεωρεί ότι το timestamp είναι
    ΗΔΗ local START time ανά 15λεπτο (00:00, 00:15, ..., 23:45) και
    ΔΕΝ το μετακινεί -15'.

    Επιστρέφει: TIMESTAMP (local START), DAM Price (€/MWh), dup_idx
    μόνο για τον ζητούμενο μήνα (YYYY-MM), και μόνο από 2025-10-01 και μετά.
    """
    fn = "load_dam_prices_15min"
    try:
        header_line = _find_header_line(dam_csv_path)
        dam = pd.read_csv(dam_csv_path, sep=None, engine="python", encoding="utf-8-sig", header=header_line)
        dam = dam.loc[:, ~dam.columns.astype(str).str.fullmatch(r"Unnamed: \d+")]
        dam.columns = [str(c).strip() for c in dam.columns]

        ts_col, price_col = _infer_dam_columns(dam)

        ts_aware = pd.to_datetime(dam[ts_col], errors="coerce", utc=True)
        if ts_aware.isna().all():
            start_local = pd.to_datetime(dam[ts_col], errors="coerce")  # naive local
        else:
            start_local = ts_aware.dt.tz_convert("Europe/Athens").dt.tz_localize(None)

        price = pd.to_numeric(
            dam[price_col].astype(str).str.replace(",", ".", regex=False),
            errors="coerce"
        )

        out = pd.DataFrame({
            "TIMESTAMP": start_local,
            "DAM Price (€/MWh)": price
        }).dropna(subset=["TIMESTAMP"])

        lb = pd.Timestamp("2025-10-01 00:00")
        out = out[out["TIMESTAMP"] >= lb]
        out = out[out["TIMESTAMP"].dt.strftime("%Y-%m") == month].copy()

        # ΔΕΝ κάνουμε sort: κρατάμε τη σειρά αρχείου, αλλά βάζουμε dup_idx
        out["dup_idx"] = out.groupby("TIMESTAMP").cumcount()

        log(fn, f"DAM 15' rows after filters: {len(out)} for {month}")
        return out.reset_index(drop=True)

    except Exception as e:
        log(fn, f"ERROR {e}")
        return None


# =================== Production reading (μήνιαία λογική) ===================

def read_production_data(file_path):
    fn = "read_production_data"
    try:
        df = pd.read_csv(file_path, sep=None, engine="python", encoding='utf-8-sig')
        log(fn, f"read {file_path}: {len(df)} rows")
        return df
    except Exception as e:
        log(fn, f"ERROR {file_path}: {e}")
        return None


# =================== ΜΗΝΙΑΙΟΣ ΥΠΟΛΟΓΙΣΜΟΣ (copied) ===================

def calculate_daily_summary_quarterly(df_prod, df_dam_15m, producer_row, month):
    """
    ΑΚΡΙΒΩΣ η ίδια λογική με το μηνιαίο ADMIE_MERGED_TIMOLOGIA:

      - Prod(END): D 00:15..23:45 + (D+1) 00:00
      - DAM(START): D 00:00..23:45
      - P[i] ↔ DAM[i] χωρίς shift.

    1/10: πετάμε τις 00:15/00:30/00:45
    26/10: κρατάμε μόνο την πρώτη εμφάνιση στα διπλά 03:00–04:00 (prod & DAM).
    """
    fn = "calculate_daily_summary_15m_by_index"
    try:
        month_str = month

        # =============== 1. ΠΑΡΑΓΩΓΗ (END TS) ================
        prod = df_prod.copy()
        prod['END_TS'] = pd.to_datetime(
            prod['TIMESTAMP'],
            format="%d/%m/%Y %H:%M",
            errors='coerce',
            dayfirst=True
        )
        prod = prod.dropna(subset=['END_TS'])
        prod = prod[prod['END_TS'].dt.strftime("%Y-%m") == month_str].copy()
        if prod.empty:
            log(fn, f"no production rows for {month_str}")
            return None, None

        prod['ΕΝΕΡΓΕΙΑ (kWh)'] = pd.to_numeric(
            prod['ΕΝΕΡΓΕΙΑ (kWh)'], errors='coerce'
        ).fillna(0.0)
        prod = prod.sort_values('END_TS').reset_index(drop=True)

        # =============== 2. DAM (START TS) ================
        dam = df_dam_15m.copy()
        dam['START_TS'] = pd.to_datetime(dam['TIMESTAMP'], errors='coerce')
        dam = dam.dropna(subset=['START_TS'])
        dam = dam[dam['START_TS'].dt.strftime("%Y-%m") == month_str].copy()
        if dam.empty:
            log(fn, f"no DAM rows for {month_str}")
            return None, None

        dam = dam.sort_values('START_TS').reset_index(drop=True)

        # =============== 3. Λίστα ημερών ================
        prod['day'] = prod['END_TS'].dt.date
        days = sorted({d for d in prod['day'] if str(d).startswith(month_str)})
        if not days:
            log(fn, f"no days in production for {month_str}")
            return None, None

        all_quarters = []

        for D in days:
            D_ts = pd.Timestamp(str(D))
            D_next = D_ts + pd.Timedelta(days=1)

            # ---- Prod: (D,00:00]..(D+1,00:00] => 00:15..23:45 + next 00:00 ----
            day_prod = prod[
                (prod['END_TS'] > D_ts) & (prod['END_TS'] <= D_next)
            ].copy()

            # 01/10/2025: πετάμε END 00:15/00:30/00:45
            if D == pd.Timestamp("2025-10-01").date():
                mask_skip = (
                    (day_prod['END_TS'].dt.date == D) & 
                    (day_prod['END_TS'].dt.hour == 0) &
                    (day_prod['END_TS'].dt.minute.isin([15, 30, 45]))
                )
                day_prod = day_prod[~mask_skip].copy()

            # 26/10/2025: intervals 03:00–04:00 → END 03:15,03:30,03:45,04:00 (κρατάμε πρώτη)
            if D == pd.Timestamp("2025-10-26").date():
                mask_win = (
                    ((day_prod['END_TS'].dt.hour == 3) & day_prod['END_TS'].dt.minute.isin([15, 30, 45])) |
                    ((day_prod['END_TS'].dt.hour == 4) & (day_prod['END_TS'].dt.minute == 0))
                )
                dup = day_prod[mask_win].duplicated(subset=['END_TS'], keep='first')
                day_prod = day_prod.drop(index=day_prod[mask_win].loc[dup].index)

            day_prod = day_prod.sort_values('END_TS').reset_index(drop=True)
            if day_prod.empty:
                continue

            # ---- DAM: D 00:00..23:45 ----
            day_dam = dam[
                (dam['START_TS'] >= D_ts) &
                (dam['START_TS'] <= D_ts + pd.Timedelta(hours=23, minutes=45))
            ].copy()

            # 26/10/2025 DAM: διπλά 03:00–03:45 → κρατάμε πρώτη εμφάνιση
            if D == pd.Timestamp("2025-10-26").date():
                mask_dam_win = (
                    (day_dam['START_TS'].dt.hour == 3) &
                    (day_dam['START_TS'].dt.minute.isin([0, 15, 30, 45]))
                )
                dup_dam = day_dam[mask_dam_win].duplicated(subset=['START_TS'], keep='first')
                day_dam = day_dam.drop(index=day_dam[mask_dam_win].loc[dup_dam].index)

            day_dam = day_dam.sort_values('START_TS').reset_index(drop=True)
            if day_dam.empty:
                continue

            # ---- P[i] ↔ DAM[i] ----
            n_p = len(day_prod)
            n_d = len(day_dam)
            n = min(n_p, n_d)
            if n == 0:
                continue
            if n_p != n_d:
                log(fn, f"Length mismatch {D}: prod={n_p}, dam={n_d}, using first {n}")

            day_prod = day_prod.iloc[:n].copy()
            day_dam = day_dam.iloc[:n].copy()

            kwh = day_prod['ΕΝΕΡΓΕΙΑ (kWh)'].to_numpy()
            price = day_dam['DAM Price (€/MWh)'].to_numpy()
            value_eur = (kwh * price) / 1000.0

            per_quarter = pd.DataFrame({
                'Περίοδος εκκαθάρισης': [D] * n,
                'ΕΝΕΡΓΕΙΑ (kWh)': kwh,
                'Αξία ενέργειας βάσει μετρήσεων (€)': value_eur
            })
            all_quarters.append(per_quarter)

        if not all_quarters:
            log(fn, f"no quarter rows after pairing for {month_str}")
            return None, None

        df_all = pd.concat(all_quarters, ignore_index=True)

        # =============== 4. ΗΜΕΡΗΣΙΑ ΣΥΝΟΛΑ ================
        df_daily = df_all.groupby('Περίοδος εκκαθάρισης', as_index=False).agg({
            'ΕΝΕΡΓΕΙΑ (kWh)': 'sum',
            'Αξία ενέργειας βάσει μετρήσεων (€)': 'sum'
        })

        df_daily['Περίοδος εκκαθάρισης'] = pd.to_datetime(
            df_daily['Περίοδος εκκαθάρισης']
        ).dt.strftime('%d/%m/%y')

        rate = float(producer_row['Μοναδιαία Χρέωση ΦοΣΕ'].values[0])
        df_daily['Προμήθεια GREEN VALUE (€)'] = (
            df_daily['ΕΝΕΡΓΕΙΑ (kWh)'] / 1000.0 * rate
        ).round(2)

        df_daily['Μεσοσταθμική Τιμή Αγοράς κατά τις ώρες παραγωγής του σταθμού'] = df_daily.apply(
            lambda row: 0 if row['ΕΝΕΡΓΕΙΑ (kWh)'] == 0 else round(
                (row['Αξία ενέργειας βάσει μετρήσεων (€)'] / row['ΕΝΕΡΓΕΙΑ (kWh)']) * 1000.0, 2
            ),
            axis=1
        )

        sum_energy = round(float(df_daily['ΕΝΕΡΓΕΙΑ (kWh)'].sum()), 2)
        sum_value  = round(float(df_daily['Αξία ενέργειας βάσει μετρήσεων (€)'].sum()), 2)
        sum_prov   = round(float(df_daily['Προμήθεια GREEN VALUE (€)'].sum()), 2)

        summary_row = pd.DataFrame([{
            'Περίοδος εκκαθάρισης': 'Σύνολο',
            'ΕΝΕΡΓΕΙΑ (kWh)': sum_energy,
            'Αξία ενέργειας βάσει μετρήσεων (€)': sum_value,
            'Προμήθεια GREEN VALUE (€)': sum_prov,
            'Μεσοσταθμική Τιμή Αγοράς κατά τις ώρες παραγωγής του σταθμού': (
                round((sum_value / sum_energy) * 1000.0, 2) if sum_energy > 0 else 0
            )
        }])

        df_final = pd.concat([df_daily, summary_row], ignore_index=True)
        log(fn, f"daily rows={len(df_final)} for month={month_str}")
        return df_final, (sum_energy, sum_value, sum_prov)

    except Exception as e:
        log(fn, f"ERROR {e}")
        return None, None


# =================== WEEKLY SUMMARY (filter πάνω στο μηνιαίο) ===================

def calculate_weekly_summary_from_month(df_prod, df_dam_month, producer_row, month_str, start_date, end_date):
    """
    1) Τρέχει calculate_daily_summary_quarterly για ΟΛΟ τον μήνα.
    2) Πετάει τη γραμμή 'Σύνολο'.
    3) Κρατά μόνο τις μέρες [start_date .. end_date].
    4) Ξαναϋπολογίζει τα σύνολα για αυτές τις μέρες.
    """
    fn = "calculate_weekly_summary_from_month"
    try:
        df_month, summary_month = calculate_daily_summary_quarterly(
            df_prod, df_dam_month, producer_row, month_str
        )
        if df_month is None or df_month.empty:
            log(fn, f"no monthly summary for {month_str}")
            return None, None

        df_no_total = df_month[df_month['Περίοδος εκκαθάρισης'] != 'Σύνολο'].copy()

        df_no_total['date_obj'] = pd.to_datetime(
            df_no_total['Περίοδος εκκαθάρισης'],
            format='%d/%m/%y',
            errors='coerce'
        )
        mask = (df_no_total['date_obj'] >= start_date) & (df_no_total['date_obj'] <= end_date)
        df_week = df_no_total[mask].copy()
        if df_week.empty:
            log(fn, f"no rows in requested week {start_date}..{end_date}")
            return None, None

        sum_energy = round(float(df_week['ΕΝΕΡΓΕΙΑ (kWh)'].sum()), 2)
        sum_value  = round(float(df_week['Αξία ενέργειας βάσει μετρήσεων (€)'].sum()), 2)
        sum_prov   = round(float(df_week['Προμήθεια GREEN VALUE (€)'].sum()), 2)

        summary_row = pd.DataFrame([{
            'Περίοδος εκκαθάρισης': 'Σύνολο Εβδομάδας',
            'ΕΝΕΡΓΕΙΑ (kWh)': sum_energy,
            'Αξία ενέργειας βάσει μετρήσεων (€)': sum_value,
            'Προμήθεια GREEN VALUE (€)': sum_prov,
            'Μεσοσταθμική Τιμή Αγοράς κατά τις ώρες παραγωγής του σταθμού': (
                round((sum_value / sum_energy) * 1000.0, 2) if sum_energy > 0 else 0
            )
        }])

        df_week = df_week.drop(columns=['date_obj'])
        df_final = pd.concat([df_week, summary_row], ignore_index=True)
        log(fn, f"weekly rows={len(df_final)}")
        return df_final, (sum_energy, sum_value, sum_prov)

    except Exception as e:
        log(fn, f"ERROR {e}")
        return None, None


# =================== Excel + PDF (weekly) ===================

def _rect_bounds(coord: str):
    if ":" not in coord:
        c, r = coordinate_to_tuple(coord)
        return (c, r, c, r)
    a, b = coord.split(":")
    c1, r1 = coordinate_to_tuple(a)
    c2, r2 = coordinate_to_tuple(b)
    return (min(c1, c2), min(r1, r2), max(c1, c2), max(r1, r2))

def _ranges_overlap(a_bounds, b_bounds):
    aL,aT,aR,aB = a_bounds
    bL,bT,bR,bB = b_bounds
    return not (aR < bL or bR < aL or aB < bT or bB < aT)

def _unmerge_in_rect(ws: Worksheet, rect: str):
    target = _rect_bounds(rect)
    to_unmerge = []
    for rg in list(ws.merged_cells.ranges):
        if _ranges_overlap(target, rg.bounds):
            to_unmerge.append(rg.coord)
    for coord in to_unmerge:
        try:
            ws.unmerge_cells(coord)
        except Exception as e:
            log("_unmerge_in_rect", f"Failed to unmerge {coord}: {e}")

def set_cell_value(ws: Worksheet, coord: str, value):
    """
    Safely set a cell value, handling merged cells by unmerging first.
    """
    # Try to unmerge any range containing this cell
    merged_ranges_to_unmerge = []
    for mrange in list(ws.merged_cells.ranges):
        if coord in mrange:
            merged_ranges_to_unmerge.append(mrange.coord)
    
    for mrange_coord in merged_ranges_to_unmerge:
        try:
            ws.unmerge_cells(mrange_coord)
        except Exception as e:
            log("set_cell_value", f"Failed to unmerge {mrange_coord}: {e}")
    
    # Now try to set the value, retrying if still MergedCell
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            cell = ws[coord]
            if not isinstance(cell, MergedCell):
                cell.value = value
                return
            # If still merged, try harder to unmerge
            for mrange in list(ws.merged_cells.ranges):
                if coord in mrange:
                    ws.unmerge_cells(mrange.coord)
        except Exception as e:
            if attempt == max_attempts - 1:
                log("set_cell_value", f"Failed to set value at {coord} after {max_attempts} attempts: {e}")
            time.sleep(0.01)
    
    # Last resort: try to write to the cell directly
    try:
        ws[coord].value = value
    except Exception as e:
        log("set_cell_value", f"Final attempt failed for {coord}: {e}")

def set_cell_property(ws: Worksheet, coord: str, prop_name: str, prop_value):
    """
    Safely set a cell property (font, alignment, etc.), handling merged cells.
    For merged cells, accesses the top-left cell of the merge.
    """
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        # Find the top-left cell of the merged range
        for mrange in ws.merged_cells.ranges:
            if coord in mrange:
                min_col, min_row = mrange.bounds[0], mrange.bounds[1]
                cell = ws.cell(row=min_row, column=min_col)
                break
    
    try:
        setattr(cell, prop_name, prop_value)
    except Exception as e:
        log("set_cell_property", f"Failed to set {prop_name} on {coord}: {e}")

def _safe_merge(ws, coord: str):
    try:
        try:
            ws.unmerge_cells(coord)
        except Exception:
            pass
        ws.merge_cells(coord)
        return True
    except Exception as e:
        log("generate_invoice_excel_weekly", f"merge {coord} failed: {e}")
        return False

def _add_logo_if_available(ws, base_dir: Path):
    try:
        existing = getattr(ws, "_images", [])
        if existing and len(existing) > 0:
            log("generate_invoice_excel_weekly", "Template already has a logo; skipping extra logo.")
            return
    except Exception:
        pass

    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception as e:
        log("generate_invoice_excel_weekly", f"Pillow/openpyxl image support missing: {e}")
        return

    logo_path = base_dir / "logo.png"
    if not logo_path.exists():
        log("generate_invoice_excel_weekly", f"logo not found: {logo_path}")
        return

    try:
        img = XLImage(str(logo_path))
        img.anchor = "A1"
        img.width  = 120
        img.height = 60
        ws.add_image(img)
        log("generate_invoice_excel_weekly", f"logo added (scaled) from {logo_path}")
    except Exception as e:
        log("generate_invoice_excel_weekly", f"logo add failed: {e}")

def make_week_dirs(start_date: pd.Timestamp, end_date: pd.Timestamp):
    iso_year, iso_week, _ = start_date.isocalendar()
    tag = f"{iso_year}-W{iso_week:02d}"
    root    = BASE_DIR / "ΕΝΗΜΕΡΩΤΙΚΑ_ΣΗΜΕΙΩΜΑΤΑ_ΕΒΔΟΜΑΔΙΑΙΑ" / tag
    xlsx_dir = root / "XLSX"
    pdf_dir  = root / "PDF"
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)
    return tag, root, xlsx_dir, pdf_dir

def determine_pdf_subfolder_name(email, email_to_companies):
    companies = email_to_companies.get(email, [])
    return clipped_folder_name([], companies, limit=MAX_FOLDER_CHARS)

def generate_invoice_excel_weekly(df_daily, summary, producer_row, start_date, end_date, xlsx_output_dir: Path, tag: str):
    fn = "generate_invoice_excel_weekly"
    try:
        if not TEMPLATE_FILE.exists():
            raise FileNotFoundError(f"Λείπει template: {TEMPLATE_FILE}")

        company_name = str(producer_row['Εταιρεία'].values[0])
        email_value  = str(producer_row.get('Email', "")) if 'Email' in producer_row else ''
        iban         = producer_row['IBAN'].values[0] if 'IBAN' in producer_row else ''
        rate         = float(producer_row['Μοναδιαία Χρέωση ΦοΣΕ'].values[0])
        sum_energy, sum_value, sum_prov = summary

        out_name  = xlsx_filename_weekly(company_name, tag)
        xlsx_path = xlsx_output_dir / out_name
        xlsx_output_dir.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active

        try:
            ws.print_area = 'A1:H55'
        except Exception:
            pass
        try:
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass

        # ⚠️ Δεν πειράζουμε heights γραμμών — μένουν όπως είναι στο template

        _add_logo_if_available(ws, BASE_DIR)

        from openpyxl.styles import Font, Alignment

        # === Header: όλα merged μαζί (D1:F2) και όλα στο κέντρο ===
        _safe_merge(ws, "D1:F2")
        ws["D1"].value = (
            "Ενημερωτικό Σημείωμα Εβδομάδας\n"
            f"{start_date.strftime('%d/%m/%y')} – {end_date.strftime('%d/%m/%y')}"
        )
        ws["D1"].font = Font(bold=True, size=11)
        ws["D1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # === Merge G1:H2 και κεντράρισμα ===
        _safe_merge(ws, "G1:H1")
        ws["G1"].value = (
            'Φορέας Σωρευτικής Εκπροσώπησης ΑΠΕ (Φο.Σ.Ε.)\n'
            'Διεύθυνση: Φιλοπάππου 19, Αθήνα 11741, Ελλάδα\n'
            'ΑΦΜ: 801961185\n'
            'ΓΕΜΗ: 167104201000\n' \
            'ΔΟΥ:ΦΑΕ Αθηνών\n' \
            'Email: info@greenvalue.gr'
        )
        ws["G1"].font = Font(bold=True, size=10)
        ws["G1"].alignment = Alignment(
            horizontal="right",
            vertical="center",
            wrap_text=True
        )
        # set_cell_value(
        #     ws, "G1",
        #     'Φορέας Σωρευτικής Εκπροσώπησης ΑΠΕ (Φο.Σ.Ε.)\n'
        #     'Διεύθυνση: Φιλοπάππου 19, Αθήνα 11741, Ελλάδα\n'
        #     'ΑΦΜ: 801961185\n'
        #     'ΓΕΜΗ: 167104201000\n' \
        #     'ΔΟΥ:ΦΑΕ Αθηνών\n' \
        #     'Email: info@greenvalue.gr'
        # )
        # set_cell_property(ws, "G1", "font", Font(size=10))
        # set_cell_property(
        #     ws, "G1", "alignment",
        #     Alignment(
        #         wrap_text=True,
        #         horizontal="right",
        #         vertical="center"
        #     )
        # )

        needed = ['Α.Μ. ΑΠΕ','Εταιρεία','ΑΦΜ','ΔΟΥ','Διεύθυνση','Email','Τεχνολογία']
        vals   = producer_row.iloc[0][needed].tolist()

        # Place producer/park info one row lower (row 4)
        for cell_ref, val in zip(['B4','C4','D4','E4','F4','G4','H4'], vals):
            set_cell_value(ws, cell_ref, val)
        set_cell_property(ws, 'C4', 'font', Font(size=13))

        # === Merge D6:F6 και κεντράρισμα ===
        _safe_merge(ws, "D6:F6")
        ws["D6"].value = (
            f"{start_date.strftime('%d/%m/%y')} - {end_date.strftime('%d/%m/%y')}"
        )
        ws["D6"].font = Font(bold=True, size=14)
        ws["D6"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        # set_cell_value(
        #     ws, "D6:F6",
        #     f"{start_date.strftime('%d/%m/%y')}-{end_date.strftime('%d/%m/%y')}"
        # )
        # set_cell_property(
        #     ws, "D6:F6", "alignment",
        #     Alignment(horizontal="center", vertical="center", wrap_text=True)
        # )

        # Unmerge ALL merged cells in the data area (C10:G41) to avoid MergedCell errors
        for rg in list(ws.merged_cells.ranges):
            bounds = rg.bounds
            # Check if this merge overlaps with data area C10:G41
            if (bounds[1] >= 10 and bounds[1] <= 19 and bounds[0] >= 3 and bounds[0] <= 7):
                try:
                    ws.unmerge_cells(rg.coord)
                    log(fn, f"Unmerged: {rg.coord}")
                except Exception as e:
                    log(fn, f"Failed to unmerge {rg.coord}: {e}")

        for row in ws.iter_rows(min_row=10, max_row=19, min_col=3, max_col=7):
            for cell in row:
                try:
                    cell.value = None
                except Exception as e:
                    log(fn, f"Could not clear {cell.coordinate}: {e}")

        start_row = 10
        for r_idx, row_vals in enumerate(df_daily.values, start=start_row):
            for c_idx, value in enumerate(row_vals, start=3):
                try:
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if isinstance(value, (int, float)):
                        try:
                            cell.number_format = '#,##0.00'
                        except Exception:
                            pass
                except Exception as e:
                    log(fn, f"Could not set cell ({r_idx},{c_idx}): {e}")

        total_row = start_row + len(df_daily) - 1
        for col in range(3, 8):
            cell_coord = f"{chr(64 + col)}{total_row}"
            set_cell_property(ws, cell_coord, 'font', Font(bold=True, size=15))

        _safe_merge(ws, "C28:D28")
        # set_cell_value(ws, 'C28', iban)
        # set_cell_property(ws, 'C28', 'font', Font(size=14, bold=True))

        ws["C28"].value = iban
        ws["C28"].font = Font(bold=True, size=14)
        ws["C28"].alignment = Alignment(
            horizontal="right",
            vertical="center",
            wrap_text=True
        )

        

        _safe_merge(ws, "C29:D29")
        set_cell_value(ws, 'C29', (pd.Timestamp.today() + pd.Timedelta(days=5)).strftime('%d/%m/%y'))
        set_cell_value(ws, 'D21', rate)
        set_cell_value(ws, 'D22', round(sum_prov, 2))
        set_cell_property(ws, 'D43', 'number_format', '#,##0.00')

        wb.save(xlsx_path)
        log(fn, f"XLSX OK: {xlsx_path}")
        print(f"✅ XLSX → {xlsx_path}")
        return str(xlsx_path), company_name, email_value

    except Exception as e:
        log(fn, f"ERROR {e}")
        print(f"❌ Excel generation failed: {e}")
        return None, None, None


    except Exception as e:
        log(fn, f"ERROR {e}")
        print(f"❌ Excel generation failed: {e}")
        return None, None, None




# def _find_soffice_path() -> str | None:
#     p = shutil.which("soffice")
#     if p:
#         return p
#     default = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
#     if os.path.exists(default):
#         return default
#     return None

def export_to_pdf_with_excel(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    fn = "export_to_pdf_excel_weekly"
    if not _HAS_XLWINGS:
        log(fn, "xlwings not available; skip Excel export.")
        return False, "excel-not-available"
    import xlwings as xw  # type: ignore
    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(os.path.abspath(xlsx_path))
        try:
            sht = wb.sheets.active
            sht.api.PageSetup.Zoom         = False
            sht.api.PageSetup.FitToPagesWide  = 1
            sht.api.PageSetup.FitToPagesTall  = False
        except Exception as e:
            log(fn, f"PageSetup warn: {e}")

        out_pdf = os.path.abspath(pdf_path)
        Path(os.path.dirname(out_pdf)).mkdir(parents=True, exist_ok=True)

        # 1) Try ExportAsFixedFormat (Windows-style COM API)
        # export_failed = False
        # try:
        #     wb.api.ExportAsFixedFormat(0, out_pdf)
        # except Exception as e_api:
        #     log(fn, f"ExportAsFixedFormat failed: {e_api}; trying fallback...")
        #     export_failed = True

        # 2) If ExportAsFixedFormat failed, try xlwings wb.to_pdf (if available)
        try:
                if hasattr(wb, "to_pdf"):
                    wb.to_pdf(out_pdf)
                    export_failed = False
                else:
                    raise AttributeError("wb.to_pdf not available")
        except Exception as e_to_pdf:
                log(fn, f"wb.to_pdf failed: {e_to_pdf}; trying AppleScript fallback...")

                # 3) AppleScript fallback: ask Microsoft Excel (Mac) to save as PDF
                # try:
                #     applescript = (
                #         'tell application "Microsoft Excel"\n'
                #         f'    open POSIX file "{os.path.abspath(xlsx_path)}"\n'
                #         '    delay 0.5\n'
                #         '    tell workbook 1\n'
                #         f'        save workbook as filename POSIX file "{out_pdf}" file format PDF file format\n'
                #         '        close saving no\n'
                #         '    end tell\n'
                #         'end tell'
                #     )
                #     res = subprocess.run(["osascript", "-e", applescript], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=60)
                #     if res.returncode == 0:
                #         export_failed = False
                #         log(fn, "AppleScript export succeeded")
                #     else:
                #         log(fn, f"AppleScript failed rc={res.returncode}\nstdout:{res.stdout}\nstderr:{res.stderr}")
                # except Exception as e_apple:
                #     log(fn, f"AppleScript fallback failed: {e_apple}")

        wb.close()

        ok = os.path.exists(out_pdf) and os.path.getsize(out_pdf) >= 500
        if ok:
            log(fn, f"OK via Excel -> {out_pdf}")
            print(f"✅ PDF (Excel) → {out_pdf}")
            return True, "excel"
        else:
            log(fn, f"Excel produced no/empty file at: {out_pdf}")
            print(f"⚠️ Empty PDF from Excel → {out_pdf}")
            return False, "excel-empty"
    except Exception as e:
        log(fn, f"ERROR (Excel): {e}")
        print(f"❌ PDF export (Excel) failed: {e}")
        return False, f"excel-error:{e}"
    finally:
        try:
            app.quit()
        except Exception:
            pass

# def export_to_pdf_with_libreoffice(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
#     fn = "export_to_pdf_libreoffice_weekly"
#     soffice = _find_soffice_path()
#     if not soffice:
#         log(fn, "LibreOffice 'soffice' not found.")
#         return False, "lo-missing"

#     outdir = os.path.abspath(os.path.dirname(pdf_path))
#     Path(outdir).mkdir(parents=True, exist_ok=True)

#     cmd = [
#         soffice,
#         "--headless","--norestore","--nolockcheck",
#         "--convert-to","pdf",
#         "--outdir", outdir,
#         os.path.abspath(xlsx_path)
#     ]

#     try:
#         res = subprocess.run(
#             cmd,
#             stdout=subprocess.PIPE,
#             stderr=subprocess.PIPE,
#             text=True,
#             timeout=180
#         )
#         if res.returncode != 0:
#             log(fn, f"ERROR rc={res.returncode}\nstdout:\n{res.stdout}\nstderr:\n{res.stderr}")
#             print(f"❌ PDF export (LibreOffice) failed rc={res.returncode}")
#             return False, f"lo-error-rc{res.returncode}"

#         produced = os.path.join(outdir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
#         if os.path.abspath(produced) != os.path.abspath(pdf_path):
#             try:
#                 if os.path.exists(pdf_path):
#                     os.remove(pdf_path)
#                 os.replace(produced, pdf_path)
#             except Exception as e:
#                 log(fn, f"Rename error: {e}")
#                 return False, f"lo-rename-error:{e}"

#         time.sleep(0.2)
#         ok = os.path.exists(pdf_path) and os.path.getsize(pdf_path) >= 500
#         if ok:
#             log(fn, f"OK via LibreOffice -> {pdf_path}")
#             print(f"✅ PDF (LibreOffice) → {pdf_path}")
#             return True, "libreoffice"
#         else:
#             log(fn, f"LibreOffice produced no/empty file at: {pdf_path}")
#             print(f"⚠️ Empty PDF from LibreOffice → {pdf_path}")
#             return False, "lo-empty"

#     except Exception as e:
#         log(fn, f"ERROR (LibreOffice): {e}")
#         print(f"❌ PDF export (LibreOffice) exception: {e}")
#         return False, f"lo-exception:{e}"

def export_to_pdf(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    ok, how = export_to_pdf_with_excel(xlsx_path, pdf_path)
    if ok:
        return True, how


# =================== Main weekly job ===================

# def timologia_weekly(start_date_str: str, end_date_str: str):
#     # Parse dates
#     try:
#         start_date = pd.to_datetime(start_date_str).floor('D')
#         end_date   = pd.to_datetime(end_date_str).floor('D')
#     except Exception:
#         print("Μη έγκυρες ημερομηνίες. Χρήση: YYYY-MM-DD")
#         return
#     if end_date < start_date:
#         print("Το τέλος είναι πριν την αρχή.")
#         return

#     if start_date.year != end_date.year or start_date.month != end_date.month:
#         print("Προς το παρόν η εβδομάδα πρέπει να είναι μέσα στον ίδιο μήνα.")
#         print("Χώρισέ την σε δύο κλήσεις (μία για κάθε μήνα).")
#         return

#     month_str = start_date.strftime('%Y-%m')

#     # 1) Χτίζουμε/ενημερώνουμε ΠΑΡΑΓΩΓΗ από GREEN_VE6
#     # ensure_production_files(start_date, end_date)

#     # 2) Φάκελοι εξόδου
#     tag, root, xlsx_dir, pdf_dir = make_week_dirs(start_date, end_date)

#     # 3) Producers (ΣΗΘΥΑ μόνο)
#     producers_df = load_producers_sithya(PRODUCERS_XLSX)
#     print(producers_df)
#     if producers_df is None or producers_df.empty:
#         print("Δεν βρέθηκαν παραγωγοί ΣΗΘΥΑ.")
#         return
#     email_to_companies, _ = build_email_groups(producers_df)

#     # 4) DAM 15' για τον μήνα (START, 2025-10-01+)
#     if not DAM_FILE_2025.exists():
#         print(f"Λείπει DAM CSV: {DAM_FILE_2025.name}")
#         return
#     df_dam_month = load_dam_quarterly_endtime(str(DAM_FILE_2025), month_str)
#     if df_dam_month is None or df_dam_month.empty:
#         print("Αποτυχία: DAM 15' prices (empty).")
#         return

#     # 5) ΠΑΡΑΓΩΓΗ_*.csv
#     if not PROD_DIR.is_dir():
#         print(f"Λείπει φάκελος: {PROD_DIR} (και δεν μπόρεσα να τον φτιάξω από downloads)")
#         return

#     for filename in os.listdir(PROD_DIR):
#         if not (filename.startswith('ΠΑΡΑΓΩΓΗ_') and filename.endswith('.csv')):
#             continue

#         file_path = PROD_DIR / filename
#         m = re.match(r'ΠΑΡΑΓΩΓΗ_(.+)\.csv', filename)
#         if not m:
#             log("timologia_weekly", f"Bad filename: {filename}")
#             continue
#         company_key = m.group(1)

#         prod_row = producers_df[producers_df['normalized_name'] == normalize_name(company_key)]
#         if prod_row.empty:
#             continue

#         company_name = str(prod_row['Εταιρεία'].values[0])
#         print(f"\n=== Επεξεργασία {filename} ({company_name}) ===")

#         df = read_production_data(str(file_path))
#         if df is None:
#             print("  -> SKIP: read_production_data επέστρεψε None")
#             continue

#         df_week, summary = calculate_weekly_summary_from_month(
#             df, df_dam_month, prod_row, month_str, start_date, end_date
#         )
#         if df_week is None:
#             print("  -> SKIP: calculate_weekly_summary_from_month επέστρεψε None")
#             continue

#         xlsx_path, company_name, email_value = generate_invoice_excel_weekly(
#             df_week, summary, prod_row, start_date, end_date, xlsx_dir, tag
#         )
#         if not xlsx_path:
#             print("  -> SKIP: generate_invoice_excel_weekly απέτυχε")
#             continue

#         email_key  = (email_value or "NO_EMAIL").strip() or "NO_EMAIL"
#         subfolder  = determine_pdf_subfolder_name(email_key, email_to_companies)
#         target_dir = pdf_dir / subfolder[:MAX_FOLDER_CHARS]
#         target_dir.mkdir(parents=True, exist_ok=True)

#         pdf_name = pdf_filename_weekly(company_name, tag)
#         pdf_path = target_dir / pdf_name

#         ok, how = export_to_pdf(xlsx_path, str(pdf_path))
#         status   = "✅ PDF" if ok else "❌ PDF"
#         print(f"  {status} [{how}] → {pdf_path}")

#     print(f"\nΈτοιμο. Δες: {root}/XLSX και {root}/PDF")

def timologia_weekly(start_date_str: str, end_date_str: str):
    # Parse dates
    try:
        start_date = pd.to_datetime(start_date_str).floor('D')
        end_date   = pd.to_datetime(end_date_str).floor('D')
    except Exception:
        print("Μη έγκυρες ημερομηνίες. Χρήση: YYYY-MM-DD")
        return

    if end_date < start_date:
        print("Το τέλος είναι πριν την αρχή.")
        return

    if start_date.year != end_date.year or start_date.month != end_date.month:
        print("Προς το παρόν η εβδομάδα πρέπει να είναι μέσα στον ίδιο μήνα.")
        print("Χώρισέ την σε δύο κλήσεις (μία για κάθε μήνα).")
        return

    month_str = start_date.strftime('%Y-%m')

    # 1) Χτίζουμε/ενημερώνουμε ΠΑΡΑΓΩΓΗ από GREEN_VE6
    # ensure_production_files(start_date, end_date)

    # 2) Φάκελοι εξόδου
    tag, root, xlsx_dir, pdf_dir = make_week_dirs(start_date, end_date)

    # 3) Producers (ΣΗΘΥΑ μόνο)
    producers_df = load_producers_sithya(PRODUCERS_XLSX)
    if producers_df is None or producers_df.empty:
        print("Δεν βρέθηκαν παραγωγοί ΣΗΘΥΑ.")
        return

    # χρήσιμο για grouping pdf σε subfolders ανά email
    email_to_companies, _ = build_email_groups(producers_df)

    # 4) DAM 15' για τον μήνα
    if not DAM_FILE_2025.exists():
        print(f"Λείπει DAM CSV: {DAM_FILE_2025.name}")
        return

    df_dam_month = load_dam_quarterly_endtime(str(DAM_FILE_2025), month_str)
    if df_dam_month is None or df_dam_month.empty:
        print("Αποτυχία: DAM 15' prices (empty).")
        return

    # 5) Παραγωγή: χρειάζεται PROD_DIR
    if not PROD_DIR.is_dir():
        print(f"Λείπει φάκελος: {PROD_DIR} (και δεν μπόρεσα να τον φτιάξω από downloads)")
        return
    
    def safe_company_folder_name(name: str) -> str:
        # 1. Unicode normalization
        name = unicodedata.normalize("NFKC", name)

        # 2. Αντικατάσταση ΟΛΩΝ των whitespace (space, NBSP, tabs κλπ) με _
        name = re.sub(r"\s+", "_", name)

        # 3. Καθάρισμα διπλών _
        name = re.sub(r"_+", "_", name)

        # 4. Trim
        return name.strip("_")

    # ------------------------------------------------------------
    # Helper: βρες το σωστό production CSV για έναν παραγωγό
    # (πρώτα νέο path, μετά fallback σε παλιό flat path)
    # ------------------------------------------------------------
    def _find_production_file_for_producer(company_name: str):
        """
        Returns Path or None
        - New structure: PROD_DIR/{ΕΤΑΙΡΕΙΑ}/ΠΑΡΑΓΩΓΗ_{ΕΤΑΙΡΕΙΑ}.csv
        - Old structure: PROD_DIR/ΠΑΡΑΓΩΓΗ_{something}.csv (normalize match)
        """
        comp = safe_company_folder_name(company_name)

        # Fallback: old structure - try direct name
        p_old_direct = PROD_DIR / f"ΠΑΡΑΓΩΓΗ_{comp}.csv"
        if p_old_direct.exists():
            return p_old_direct

        # Fallback: scan PROD_DIR for a match by normalized name
        comp_norm = normalize_name(comp)
        try:
            for fn in os.listdir(PROD_DIR):
                if not (fn.startswith("ΠΑΡΑΓΩΓΗ_") and fn.endswith(".csv")):
                    continue
                m = re.match(r'ΠΑΡΑΓΩΓΗ_(.+)\.csv', fn)
                if not m:
                    continue
                key = m.group(1)
                if normalize_name(key) == comp_norm:
                    return PROD_DIR / fn
        except Exception:
            pass

        return None

    # ------------------------------------------------------------
    # Κύριο loop: πάνω στους παραγωγούς ΣΗΘΥΑ (producers_df)
    # ------------------------------------------------------------
    processed = 0
    skipped_no_file = 0
    skipped_errors = 0

    # Αν δεν υπάρχει normalized_name, το φτιάχνουμε πρόχειρα
    if "normalized_name" not in producers_df.columns:
        producers_df = producers_df.copy()
        producers_df["normalized_name"] = producers_df["Εταιρεία"].astype(str).apply(normalize_name)

    for _, prod_row_series in producers_df.iterrows():
        # κάνουμε prod_row DataFrame 1-γραμμής για να ταιριάζει με τις υπάρχουσες συναρτήσεις σου
        prod_row = producers_df.loc[[prod_row_series.name]]

        company_name = str(prod_row_series.get("Εταιρεία", "")).strip()
        if not company_name:
            continue

        file_path = _find_production_file_for_producer(company_name)
        if file_path is None or not file_path.exists():
            skipped_no_file += 1
            print(f"\n=== {company_name} ===")
            print("  -> SKIP: Δεν βρέθηκε αρχείο παραγωγής (ΠΑΡΑΓΩΓΗ_*.csv)")
            continue

        print(f"\n=== Επεξεργασία παραγωγού: {company_name} ===")
        print(f"  Production file: {file_path.name}")

        df = read_production_data(str(file_path))
        if df is None:
            skipped_errors += 1
            print("  -> SKIP: read_production_data επέστρεψε None")
            continue

        df_week, summary = calculate_weekly_summary_from_month(
            df, df_dam_month, prod_row, month_str, start_date, end_date
        )
        if df_week is None:
            skipped_errors += 1
            print("  -> SKIP: calculate_weekly_summary_from_month επέστρεψε None")
            continue

        xlsx_path, company_name_out, email_value = generate_invoice_excel_weekly(
            df_week, summary, prod_row, start_date, end_date, xlsx_dir, tag
        )
        if not xlsx_path:
            skipped_errors += 1
            print("  -> SKIP: generate_invoice_excel_weekly απέτυχε")
            continue

        # PDF export per email group subfolder
        email_key  = (email_value or "NO_EMAIL").strip() or "NO_EMAIL"
        subfolder  = determine_pdf_subfolder_name(email_key, email_to_companies)
        target_dir = pdf_dir / subfolder[:MAX_FOLDER_CHARS]
        target_dir.mkdir(parents=True, exist_ok=True)

        pdf_name = pdf_filename_weekly(company_name_out, tag)
        pdf_path = target_dir / pdf_name

        ok, how = export_to_pdf(xlsx_path, str(pdf_path))
        status  = "✅ PDF" if ok else "❌ PDF"
        print(f"  {status} [{how}] → {pdf_path}")

        processed += 1

    print("\n" + "=" * 70)
    print(f"Ολοκληρώθηκε.")
    print(f"Processed: {processed}")
    print(f"Skipped (no production file): {skipped_no_file}")
    print(f"Skipped (errors): {skipped_errors}")
    print(f"Δες: {root}/XLSX και {root}/PDF")


if __name__ == "__main__":
    start = input("Δώσε αρχή εβδομάδας (YYYY-MM-DD): ").strip()
    end   = input("Δώσε τέλος εβδομάδας (YYYY-MM-DD): ").strip()
    timologia_weekly(start, end)