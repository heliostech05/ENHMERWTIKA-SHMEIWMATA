#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
excel_export.py — Δημιουργία Excel ενημερωτικών σημειωμάτων και εξαγωγή σε PDF.

Περιλαμβάνει:
- Δημιουργία Excel από template (Invoice_GREEN_VALUE_01.xlsx)
- Εισαγωγή barplot ημερήσιας παραγωγής
- Εξαγωγή σε PDF μέσω xlwings
"""

import logging
import os
from calendar import monthrange
from copy import copy
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
import openpyxl.styles as xls

from .config import (
    BASE_DIR, TEMPLATE_PATH, GREEK_MONTHS_GENITIVE,
    MAX_FOLDER_CHARS, MAX_FILENAME_CHARS, OUTPUT_DIR,
)
from .helpers import sanitize_name, clipped_filename, clipped_folder_name

log = logging.getLogger("MONTHLY.excel_export")

# ===== Font registration =====
_FONTS_REGISTERED = False


def _register_fonts():
    """Εγγραφή τοπικών fonts (μία φορά) ώστε να δουλεύουν χωρίς system install."""
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    font_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "fonts")
    for fname in ["CenturyGothic.ttf", "centurygothic.ttf",
                   "GOTHICB.ttf", "centurygothic_bold.ttf",
                   "GOTHICI.ttf", "GOTHICBI.ttf"]:
        fpath = os.path.join(font_dir, fname)
        if os.path.exists(fpath):
            fm.fontManager.addfont(fpath)
    matplotlib.rcParams["font.family"] = "Century Gothic"
    matplotlib.rcParams["font.sans-serif"] = [
        "Century Gothic", "CenturyGothic", "DejaVu Sans",
    ]
    _FONTS_REGISTERED = True


# ===== Plot =====

def _add_daily_plot(ws, df_daily, anchor_cell="B56", color_hex="#22A052"):
    """Barplot ημερήσιας παραγωγής (χωρίς τη γραμμή 'Σύνολο')."""
    _register_fonts()
    try:
        df_plot = df_daily[
            df_daily["Περίοδος εκκαθάρισης"].str.contains("Σύνολο", case=False) == False
        ].copy()
        x = df_plot["Περίοδος εκκαθάρισης"].astype(str)
        y = pd.to_numeric(df_plot["ΕΝΕΡΓΕΙΑ (kWh)"], errors="coerce")

        plt.figure(figsize=(7.5, 3.0), dpi=160)
        plt.bar(x, y, color=color_hex, width=0.6)

        plt.title("Διάγραμμα Ημερήσιας Παραγωγής (kWh)",
                  fontname="Century Gothic", fontsize=13, pad=15)
        plt.xlabel("")
        plt.ylabel("kWh", fontname="Century Gothic", fontsize=11)
        plt.xticks(rotation=45, ha="right", fontsize=9, fontname="Century Gothic")
        plt.yticks(fontsize=9, fontname="Century Gothic")
        plt.grid(True, linestyle="--", alpha=0.4)
        plt.tight_layout(pad=1.0)

        plot_path = str(BASE_DIR / "daily_prod_plot.png")
        plt.savefig(plot_path, dpi=160, transparent=True)
        plt.close()

        img = XLImage(plot_path)
        img.anchor = anchor_cell
        img.width = 1100
        img.height = 400
        ws.add_image(img)

    except Exception as e:
        log.warning("Plot insertion failed: %s", e)


# ===== Header formatting =====

def fix_header_rows(ws):
    """Ρύθμιση ύψους header γραμμών στο template."""
    ws.row_dimensions[1].height = 160
    ws.row_dimensions[2].height = 80
    ws.row_dimensions[4].height = 100
    ws.row_dimensions[58].height = 80

    ws["C2"].alignment = xls.Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=False,
    )


# ===== Output folder structure =====

def make_base_dirs(month):
    """Δημιουργία φακέλων XLSX και PDF εξόδου."""
    root = os.path.join(str(OUTPUT_DIR), month)
    xlsx_dir = os.path.join(root, "XLSX")
    pdf_dir = os.path.join(root, "PDF")
    os.makedirs(xlsx_dir, exist_ok=True)
    os.makedirs(pdf_dir, exist_ok=True)
    return root, xlsx_dir, pdf_dir


def determine_pdf_subfolder_name(email, email_to_companies, email_to_customs):
    """Δημιουργία ονόματος subfolder για PDF ανά email ομάδα."""
    customs = email_to_customs.get(email, [])
    companies = email_to_companies.get(email, [])
    return clipped_folder_name(customs, companies, limit=MAX_FOLDER_CHARS)


def xlsx_filename(company_name, month):
    return clipped_filename(company_name, month, "xlsx", limit=MAX_FILENAME_CHARS)


def pdf_filename(company_name, month):
    return clipped_filename(company_name, month, "pdf", limit=MAX_FILENAME_CHARS)


# ===== Excel generation =====

def generate_invoice_excel(df_daily_energy, summary, producer_row, month, xlsx_output_dir):
    """
    Δημιουργία Excel ενημερωτικού σημειώματος από template.

    Γεμίζει τα στοιχεία εταιρείας, τα ημερήσια δεδομένα ενέργειας,
    τα σύνολα, και εισάγει barplot παραγωγής.
    """
    _register_fonts()
    try:
        year, month_number = month.split('-')
        year_int = int(year)
        month_int = int(month_number)

        if not TEMPLATE_PATH.exists():
            raise FileNotFoundError(f"Λείπει template: {TEMPLATE_PATH}")

        company_name = str(producer_row['Εταιρεία'].values[0])
        email_value = str(producer_row['Email'].values[0]) if 'Email' in producer_row else ''
        iban = producer_row['IBAN'].values[0] if 'IBAN' in producer_row else ''
        rate = float(producer_row['Μοναδιαία Χρέωση ΦοΣΕ'].values[0])
        sum_energy, sum_value, sum_prov = summary

        out_name = xlsx_filename(company_name, month)
        xlsx_path = os.path.join(xlsx_output_dir, out_name)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        wb = load_workbook(str(TEMPLATE_PATH))
        ws = wb.active

        ws.row_dimensions[1].height = 160
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 40
        ws.row_dimensions[5].height = 50

        for rng in ('C2:G2',):
            try:
                ws.unmerge_cells(rng)
            except Exception:
                pass

        for col in range(1, 9):
            c = ws.cell(row=2, column=col)
            if not isinstance(c, MergedCell):
                c.value = None

        ws.merge_cells('C2:G2')
        ws['C2'] = f'Ενημερωτικό Σημείωμα {GREEK_MONTHS_GENITIVE.get(month_number, "")} {year}'
        ws['C2'].font = xls.Font(name="Century Gothic", bold=True, underline='single', size=20)
        ws['C2'].alignment = xls.Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('G1:H1')
        ws['G1'] = (
            'Φορέας Σωρευτικής Εκπροσώπησης ΑΠΕ (Φο.Σ.Ε.)\n'
            'Διεύθυνση: Φιλοπάππου 19, Αθήνα 11741, Ελλάδα \n'
            'ΑΦΜ: 801961185, ΓΕΜΗ: 167104201000, ΔΟΥ:ΦΑΕ Αθηνών \n'
            'e-mail: info@greenvalue.gr'
        )
        ws['G1'].font = xls.Font(name="Century Gothic", size=14)
        ws['G1'].alignment = xls.Alignment(wrap_text=True, vertical='top')

        ws.row_dimensions[7].height = 14

        needed = ['Α.Μ. ΑΠΕ', 'Εταιρεία', 'ΑΦΜ', 'ΔΟΥ', 'Διεύθυνση', 'Email', 'Τεχνολογία']
        vals = producer_row.iloc[0][needed].tolist()
        for cell_ref, val in zip(['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4'], vals):
            ws[cell_ref] = val
        ws['C4'].font = xls.Font(name="Century Gothic", size=13)

        today = pd.Timestamp.today().strftime('%d/%m/%y')
        start_date = pd.to_datetime(df_daily_energy.iloc[0]['Περίοδος εκκαθάρισης'], dayfirst=True)
        end_date = pd.to_datetime(df_daily_energy.iloc[-2]['Περίοδος εκκαθάρισης'], dayfirst=True)
        ws['B6'] = today
        ws['C6'] = 'Αρχική'
        ws.merge_cells('D6:F6')
        ws['D6'] = f"{start_date.strftime('%d/%m/%y')}-{end_date.strftime('%d/%m/%y')}"

        for row in ws.iter_rows(min_row=10, max_row=41, min_col=3, max_col=7):
            for cell in row:
                cell.value = None

        start_row = 10
        for r_idx, row_vals in enumerate(df_daily_energy.values, start=start_row):
            is_total_row = str(row_vals[0]).strip().casefold().startswith("σύνολο")
            for c_idx, value in enumerate(row_vals, start=3):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                if is_total_row:
                    total_font = copy(cell.font)
                    total_font.bold = True
                    cell.font = total_font

        first_data_row = 10
        last_data_row = first_data_row + len(df_daily_energy) - 1
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(row=r, column=7).alignment = xls.Alignment(
                horizontal='center', vertical='center'
            )

        ws['H45'] = round(sum_energy / 1000, 2)
        ws['H45'].number_format = '#,##0.00'
        ws['H46'] = round(sum_value, 2)
        ws['H46'].number_format = '#,##0.00'
        ws['C52'] = iban
        ws['C52'].font = xls.Font(name="Century Gothic", size=14, bold=True)
        ws['C53'] = (pd.Timestamp.today() + pd.Timedelta(days=2)).strftime('%d/%m/%y')
        ws['D45'] = rate
        ws['D46'] = round(sum_prov, 2)
        ws['D46'].number_format = '#,##0.00'

        # Το template είναι για 31 ημέρες. Για μικρότερους μήνες, κρύβουμε
        # σειρές από 41 και προς τα πάνω (41, 40, 39, ...) ώστε να μην
        # αλλοιωθεί η δομή των πινάκων που βρίσκονται πιο κάτω.
        days_in_month = monthrange(year_int, month_int)[1]
        rows_to_hide = max(0, 31 - days_in_month)
        for i in range(rows_to_hide):
            row_idx = 41 - i
            ws.row_dimensions[row_idx].hidden = True

        try:
            plot_row = 59
            ws.row_dimensions[58].height = 10
            _add_daily_plot(
                ws,
                df_daily_energy,
                anchor_cell=f"B{plot_row}",
                color_hex="#22A052",
            )
        except Exception as e:
            log.warning("Plot insertion failed: %s", e)

        fix_header_rows(ws)
        wb.save(xlsx_path)
        log.info("XLSX created: %s", xlsx_path)
        return xlsx_path, company_name, email_value

    except Exception as e:
        log.error("Excel generation failed for %s: %s",
                  producer_row.get('Εταιρεία', '?'), e)
        print(f"❌ Excel generation failed: {e}")
        return None, None, None


# ===== PDF export =====

def _verify_pdf(path: str, min_bytes: int = 500):
    """Ελέγχει αν το PDF δημιουργήθηκε σωστά (μέγεθος ≥ min_bytes)."""
    return os.path.exists(path) and os.path.getsize(path) >= min_bytes


def export_to_pdf_with_excel(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    """Εξαγωγή XLSX → PDF μέσω xlwings (Microsoft Excel)."""
    try:
        import xlwings as xw
    except Exception as e:
        log.debug("xlwings not available: %s", e)
        return False, "excel-not-available"

    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(os.path.abspath(xlsx_path))
        sht = wb.sheets.active
        try:
            sht.api.PageSetup.Zoom = False
            sht.api.PageSetup.FitToPagesWide = 1
            sht.api.PageSetup.PrintArea = "B1:H57"
        except Exception:
            pass

        out_pdf = os.path.abspath(pdf_path)
        Path(os.path.dirname(out_pdf)).mkdir(parents=True, exist_ok=True)

        try:
            if hasattr(wb, "to_pdf"):
                wb.to_pdf(out_pdf)
            else:
                raise AttributeError("wb.to_pdf not available")
        except Exception as e_to_pdf:
            log.warning("wb.to_pdf failed: %s", e_to_pdf)

        try:
            wb.close()
        except Exception:
            pass

        ok = _verify_pdf(out_pdf)
        if ok:
            log.debug("PDF exported via Excel: %s", out_pdf)
            return True, "excel"
        else:
            log.warning("Excel export produced no/empty PDF: %s", out_pdf)
            return False, "excel-empty"

    except Exception as e:
        log.error("PDF export failed: %s", e)
        return False, f"excel-error:{e}"
    finally:
        try:
            app.quit()
        except Exception:
            pass


def export_to_pdf(xlsx_path: str, pdf_path: str) -> tuple[bool, str]:
    """Εξαγωγή PDF — δοκιμάζει Excel (xlwings)."""
    ok, how = export_to_pdf_with_excel(xlsx_path, pdf_path)
    if ok:
        return True, how
    return False, how
